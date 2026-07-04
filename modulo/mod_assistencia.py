# =============================================================
# modulo/mod_assistencia.py
# Formulário S-88-T (registro da assistência às reuniões).
#
# ATUALIZAÇÃO: paleta recolorida para o tema claro (era azul-marinho
# escuro #2c3e50) e aceita pode_editar=True/False — quando False, os
# campos ficam desabilitados e o botão de salvar some.
#
# IMPORTANTE: este módulo usa sua PRÓPRIA coleção no Firestore
# (congregacoes/{congregacao_id}/assistencia), diferente da
# coleção "assistencia_reunioes" usada em database.py — isso já
# era assim no arquivo original e foi mantido exatamente igual.
# =============================================================
import io
import streamlit as st
from datetime import datetime

MESES_ORDEM = [
    "Setembro", "Outubro", "Novembro", "Dezembro",
    "Janeiro", "Fevereiro", "Março", "Abril",
    "Maio", "Junho", "Julho", "Agosto",
]

TIPOS = ["Reunião do Meio de Semana", "Reunião do Fim de Semana"]

_COR_DOURADO = "#C9A227"
_COR_TEXTO   = "#6B5E3C"

# ─────────────────────────────────────────────────────────────
# FIRESTORE helpers
# ─────────────────────────────────────────────────────────────

def _col(db, cong_id: str):
    return db.collection("congregacoes").document(cong_id).collection("assistencia")


def _doc_id(tipo: str, ano_ref: str) -> str:
    slug = tipo.lower().replace(" ", "_").replace("ã", "a").replace("é", "e")
    return f"{slug}_{ano_ref.replace('/', '-')}"


def _carregar(db, cong_id: str, tipo: str, ano_ref: str) -> dict:
    try:
        doc = _col(db, cong_id).document(_doc_id(tipo, ano_ref)).get()
        if doc.exists:
            data = doc.to_dict() or {}
            return data.get("meses", {})
    except Exception:
        pass
    return {}


def _salvar(db, cong_id: str, tipo: str, ano_ref: str, meses: dict) -> bool:
    try:
        _col(db, cong_id).document(_doc_id(tipo, ano_ref)).set({
            "tipo_reuniao": tipo,
            "ano_referencia": ano_ref,
            "meses": meses,
            "atualizado_em": datetime.utcnow().isoformat(),
        })
        return True
    except Exception as exc:
        st.error(f"Erro ao salvar: {exc}")
        return False


# ─────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────

def _gerar_excel(tipo: str, ano_ref: str, meses_data: dict) -> bytes | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Assistência"

        borda = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        cinza   = PatternFill("solid", fgColor="F1EAD2")
        dourado = PatternFill("solid", fgColor="C9A84C")
        creme   = PatternFill("solid", fgColor="FFFDF6")

        ws.merge_cells("A1:D1")
        ws["A1"] = "REGISTRO DA ASSISTÊNCIA ÀS REUNIÕES CONGREGACIONAIS"
        ws["A1"].font = Font(bold=True, size=13, color="6B5E3C")
        ws["A1"].fill = creme
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:D2")
        ws["A2"] = f"{tipo.upper()}   —   ANO DE SERVIÇO: {ano_ref}"
        ws["A2"].font = Font(bold=True, size=11, color="8A6D14")
        ws["A2"].fill = creme
        ws["A2"].alignment = Alignment(horizontal="center")

        headers = ["Mês", "Qtd. Reuniões", "Assistência Total", "Média por Semana"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(bold=True, size=10)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.fill = cinza
            c.border = borda

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.row_dimensions[3].height = 30

        soma_qtd = soma_tot = meses_com_dado = 0

        for i, mes in enumerate(MESES_ORDEM, 4):
            dados = meses_data.get(mes, {})
            qtd   = dados.get("qtd", 0) or 0
            total = dados.get("total", 0) or 0
            media = round(total / qtd, 1) if qtd > 0 else 0

            valores = [mes, qtd or "", total or "", media or ""]
            for col, v in enumerate(valores, 1):
                c = ws.cell(row=i, column=col, value=v)
                c.alignment = Alignment(horizontal="left" if col == 1 else "center")
                c.border = borda

            soma_qtd += qtd
            soma_tot += total
            if qtd > 0:
                meses_com_dado += 1

        media_geral = round(soma_tot / meses_com_dado, 1) if meses_com_dado > 0 else 0
        row_tot = 4 + len(MESES_ORDEM)
        totais = ["Assistência média por mês", soma_qtd, soma_tot, media_geral]
        for col, v in enumerate(totais, 1):
            c = ws.cell(row=row_tot, column=col, value=v)
            c.font = Font(bold=True, color="1A1A1A")
            c.fill = dourado
            c.alignment = Alignment(horizontal="left" if col == 1 else "center")
            c.border = borda

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        return None


# ─────────────────────────────────────────────────────────────
# CSS injetado uma única vez
# ─────────────────────────────────────────────────────────────

def _inject_css():
    st.markdown(f"""
<style>
.s88-header {{
    background: #FBF1D4;
    color: {_COR_TEXTO};
    padding: 10px 18px 6px;
    border-radius: 8px 8px 0 0;
    margin-bottom: 0;
    border: 1px solid #E9D48E;
    border-bottom: none;
}}
.s88-header h3 {{
    margin: 0 0 2px;
    font-size: 1.05rem;
    font-weight: 700;
    letter-spacing: .04em;
    color: #8A6D14;
}}
.s88-header p  {{
    margin: 0;
    font-size: .78rem;
    color: {_COR_TEXTO};
}}

.s88-table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 0.82rem;
}}
.s88-table th {{
    background: #FBF1D4;
    color: #8A6D14;
    text-align: center;
    padding: 7px 4px;
    font-size: 0.78rem;
    font-weight: 700;
    border: 1px solid #E9D48E;
}}
.s88-table td {{
    border: 1px solid #EEE3C7;
    padding: 3px 6px;
    vertical-align: middle;
    text-align: center;
}}
.s88-table td.mes-label {{
    text-align: left;
    font-weight: 500;
    color: #1A1A1A;
    padding-left: 10px;
    white-space: nowrap;
}}
.s88-table tr.totais-row td {{
    background: {_COR_DOURADO};
    color: #1A1A1A;
    font-weight: 700;
    font-size: 0.8rem;
}}

div[data-testid="stNumberInput"] {{ margin: 0 !important; }}
div[data-testid="stNumberInput"] input {{
    padding: 3px 6px !important;
    font-size: 0.82rem !important;
    text-align: center !important;
}}
div[data-testid="stNumberInput"] > div > div:last-child {{ display: none !important; }}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# Bloco de um tipo de reunião (meio / fim de semana)
# ─────────────────────────────────────────────────────────────

def _bloco_reuniao(db, cong_id: str, tipo: str, ano_ref: str, prefixo: str, pode_editar: bool):
    cache_key = f"assis_{prefixo}_{ano_ref}"
    if cache_key not in st.session_state:
        st.session_state[cache_key] = _carregar(db, cong_id, tipo, ano_ref)

    meses_data: dict = st.session_state[cache_key]

    st.markdown(f"""
<div class="s88-header">
  <h3>{tipo.upper()}</h3>
  <p>Ano de serviço: <strong style="color:#1A1A1A">{ano_ref}</strong></p>
</div>
""", unsafe_allow_html=True)

    st.markdown("""
<table class="s88-table">
  <thead>
    <tr>
      <th style="width:22%">Mês</th>
      <th style="width:26%">Qtd. de Reuniões</th>
      <th style="width:26%">Assistência Total</th>
      <th style="width:26%">Média por Semana</th>
    </tr>
  </thead>
</table>
""", unsafe_allow_html=True)

    soma_qtd = soma_tot = meses_com_dado = 0
    medias = []

    for mes in MESES_ORDEM:
        dados_mes  = meses_data.get(mes, {})
        val_qtd    = int(dados_mes.get("qtd",   0) or 0)
        val_total  = int(dados_mes.get("total", 0) or 0)

        col_mes, col_qtd, col_tot, col_med = st.columns([2.2, 2.6, 2.6, 2.6])

        with col_mes:
            st.markdown(f"""
<div style="
    padding: 6px 10px; font-weight: 500; font-size: 0.83rem;
    color: #1A1A1A; border-bottom: 1px solid #EEE3C7;
    height: 44px; display: flex; align-items: center;
">{mes}</div>""", unsafe_allow_html=True)

        with col_qtd:
            qtd = st.number_input(
                "qtd", min_value=0, max_value=9999, value=val_qtd, step=1,
                label_visibility="collapsed", key=f"{prefixo}_{mes}_qtd",
                disabled=not pode_editar,
            )

        with col_tot:
            total = st.number_input(
                "tot", min_value=0, max_value=99999, value=val_total, step=1,
                label_visibility="collapsed", key=f"{prefixo}_{mes}_tot",
                disabled=not pode_editar,
            )

        media = round(total / qtd, 1) if qtd > 0 else 0.0
        with col_med:
            st.markdown(f"""
<div style="
    padding: 6px 10px; font-size: 0.83rem;
    color: {'#1A1A1A' if media > 0 else '#B5AC8F'};
    border-bottom: 1px solid #EEE3C7; height: 44px;
    display: flex; align-items: center; justify-content: center;
    font-weight: {'600' if media > 0 else '400'};
">{media if media > 0 else '—'}</div>""", unsafe_allow_html=True)

        meses_data[mes] = {"qtd": int(qtd), "total": int(total)}
        soma_qtd += int(qtd)
        soma_tot += int(total)
        if int(qtd) > 0:
            meses_com_dado += 1
        medias.append(media)

    st.session_state[cache_key] = meses_data

    media_geral = round(soma_tot / meses_com_dado, 1) if meses_com_dado > 0 else 0.0
    st.markdown(f"""
<table class="s88-table">
  <tbody>
    <tr class="totais-row">
      <td style="width:22%;text-align:left;padding-left:10px">Assistência média por mês</td>
      <td style="width:26%">{soma_qtd if soma_qtd else '—'}</td>
      <td style="width:26%">{soma_tot if soma_tot else '—'}</td>
      <td style="width:26%">{media_geral if media_geral else '—'}</td>
    </tr>
  </tbody>
</table>
""", unsafe_allow_html=True)

    return meses_data, soma_qtd, soma_tot, media_geral


# ─────────────────────────────────────────────────────────────
# Função principal
# ─────────────────────────────────────────────────────────────

def render_tab_assistencia(db, congregacao_id: str, pode_editar: bool = True):
    _inject_css()

    st.markdown("""
<div style="
    background:#FBF1D4; color:#8A6D14; padding:14px 20px;
    border-radius:8px; margin-bottom:20px; border:1px solid #E9D48E;
">
  <h2 style="margin:0;font-size:1.1rem;letter-spacing:.03em;color:#1A1A1A;">
    📋 REGISTRO DA ASSISTÊNCIA ÀS REUNIÕES CONGREGACIONAIS
  </h2>
  <p style="margin:4px 0 0;font-size:.78rem;color:#8A6D14">
    Formulário S-88-T · Preencha mês a mês e salve no Firestore
  </p>
</div>
""", unsafe_allow_html=True)

    if not pode_editar:
        import permissoes
        permissoes.aviso_somente_leitura()

    ano_atual  = datetime.now().year
    anos_opcao = [f"{a}/{a+1}" for a in range(ano_atual - 5, ano_atual + 2)]
    mes_atual  = datetime.now().month
    ano_padrao = f"{ano_atual}/{ano_atual+1}" if mes_atual >= 9 else f"{ano_atual-1}/{ano_atual}"

    col_sel, col_spacer = st.columns([2, 5])
    with col_sel:
        ano_ref = st.selectbox(
            "Ano de Serviço", options=anos_opcao,
            index=anos_opcao.index(ano_padrao) if ano_padrao in anos_opcao else 0,
            key="s88_ano_ref",
        )

    st.markdown("---")

    col_msem, col_gap, col_fsem = st.columns([1, 0.04, 1])

    with col_msem:
        dados_msem, qtd_msem, tot_msem, med_msem = _bloco_reuniao(
            db, congregacao_id, tipo="Reunião do Meio de Semana",
            ano_ref=ano_ref, prefixo="msem", pode_editar=pode_editar,
        )

    with col_gap:
        st.markdown(
            "<div style='border-left:2px solid #EEE3C7;height:100%;min-height:600px'></div>",
            unsafe_allow_html=True,
        )

    with col_fsem:
        dados_fsem, qtd_fsem, tot_fsem, med_fsem = _bloco_reuniao(
            db, congregacao_id, tipo="Reunião do Fim de Semana",
            ano_ref=ano_ref, prefixo="fsem", pode_editar=pode_editar,
        )

    st.markdown("---")

    if pode_editar:
        col_salvar, col_excel_m, col_excel_f, col_imprimir = st.columns([1.5, 1.5, 1.5, 1])
        with col_salvar:
            if st.button("💾 Salvar no Firestore", type="primary", use_container_width=True):
                ok_m = _salvar(db, congregacao_id, "Reunião do Meio de Semana", ano_ref, dados_msem)
                ok_f = _salvar(db, congregacao_id, "Reunião do Fim de Semana",  ano_ref, dados_fsem)
                if ok_m and ok_f:
                    st.success("✅ Dados salvos com sucesso!")
                else:
                    st.error("Falha ao salvar um ou ambos os registros.")
    else:
        col_excel_m, col_excel_f, col_imprimir = st.columns(3)

    with col_excel_m:
        bytes_m = _gerar_excel("Reunião do Meio de Semana", ano_ref, dados_msem)
        if bytes_m:
            st.download_button(
                label="📥 Excel — Meio de Semana", data=bytes_m,
                file_name=f"assistencia_meio_semana_{ano_ref.replace('/','-')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.info("openpyxl não instalado")

    with col_excel_f:
        bytes_f = _gerar_excel("Reunião do Fim de Semana", ano_ref, dados_fsem)
        if bytes_f:
            st.download_button(
                label="📥 Excel — Fim de Semana", data=bytes_f,
                file_name=f"assistencia_fim_semana_{ano_ref.replace('/','-')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.info("openpyxl não instalado")

    with col_imprimir:
        st.markdown("""
<button onclick="window.print()" style="
    width:100%; padding:8px 0; background:#FBF1D4; color:#8A6D14;
    border:1px solid #E9D48E; border-radius:6px; font-size:0.85rem;
    font-weight:600; cursor:pointer; letter-spacing:.03em;
">🖨️ Imprimir</button>
""", unsafe_allow_html=True)

    with st.expander("📊 Resumo do ano de serviço", expanded=False):
        r1, r2, r3, r4 = st.columns(4)
        r1.metric("Total reuniões (meio)",    qtd_msem)
        r2.metric("Assistência total (meio)", tot_msem)
        r3.metric("Total reuniões (fim)",     qtd_fsem)
        r4.metric("Assistência total (fim)",  tot_fsem)

        r5, r6 = st.columns(2)
        r5.metric("Média geral — meio de semana", f"{med_msem}")
        r6.metric("Média geral — fim de semana",  f"{med_fsem}")
