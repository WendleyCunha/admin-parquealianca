# =============================================================
# 1. IMPORTS
# =============================================================
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import json
import io
import zipfile
import unicodedata
import os
import base64
from datetime import datetime, date
from difflib import SequenceMatcher
from google.cloud import firestore
from google.oauth2 import service_account

from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm


# =============================================================
# 2. CONFIGURAÇÃO DA PÁGINA
# =============================================================
st.set_page_config(
    page_title="Parque Aliança · Gestão",
    layout="wide",
    page_icon="🏢",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

*, *::before, *::after { box-sizing: border-box; }

html, body, [class*="css"], .stMarkdown p {
    font-family: 'Inter', sans-serif !important;
}

.stApp {
    background: linear-gradient(180deg, #FAF7EE 0%, #F4EFDD 100%) !important;
    color: #1A1A1A !important;
}
.main .block-container {
    padding: 1.5rem 2.5rem 3rem !important;
    max-width: 1400px;
}

/* ---- Barra superior única em preto (referência: jw.org) ---- */
header[data-testid="stHeader"] {
    background: #111111 !important;
    border-bottom: 3px solid #C9A227 !important;
    height: 3.1rem !important;
}
header[data-testid="stHeader"] * {
    color: #F4EFDD !important;
}
[data-testid="stToolbar"] { color: #F4EFDD !important; }

/* ---- Sidebar clara e dourada ---- */
[data-testid="stSidebar"] {
    background: #FFFDF6 !important;
    border-right: 1px solid #EEE3B8 !important;
}

[data-testid="stSidebar"],
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] span {
    color: #6B5E3C !important;
}

[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 {
    color: #1A1A1A !important;
}

[data-testid="stSidebar"] .pa-card *,
[data-testid="stSidebar"] .pa-metric * {
    color: #1A1A1A !important;
}
[data-testid="stSidebar"] .pa-metric-label {
    color: #9C8A46 !important;
}

[data-testid="stSidebar"] [data-testid="stSelectbox"] > div > div {
    background: #FFFFFF !important;
    border: 1px solid #E7D9A0 !important;
    color: #1A1A1A !important;
    border-radius: 8px !important;
}

.sidebar-brand { text-align: center; padding: 8px 0 2px; }
.sidebar-brand-title {
    font-size: 1.05rem !important; font-weight: 800 !important;
    color: #1A1A1A !important; letter-spacing: -0.01em;
}
.sidebar-brand-sub {
    font-size: 0.7rem !important; font-weight: 700 !important;
    color: #B4952E !important; text-transform: uppercase;
    letter-spacing: 0.09em; margin-top: 2px;
}
.sidebar-divider {
    border: none !important; border-top: 1px solid #EEE3B8 !important;
    margin: 14px 0 !important;
}
.mes-badge {
    display: inline-flex; align-items: center; gap: 6px;
    background: #FBF1D4; border: 1px solid #E9D48E; border-radius: 999px;
    padding: 5px 14px; font-size: 0.75rem; font-weight: 700; color: #8A6D14;
}
.mes-dot { width: 7px; height: 7px; border-radius: 50%; background: #C9A227; display: inline-block; }

h1, h2, h3, h4, h5 {
    color: #1A1A1A !important;
    font-family: 'Inter', sans-serif !important;
}
h1 { font-size: 1.9rem !important; font-weight: 700 !important; letter-spacing: -0.02em !important; }
h2 { font-weight: 600 !important; font-size: 1.3rem !important; }

[data-testid="stTabs"] [data-testid="stTab"] {
    color: #888888 !important;
    font-weight: 500 !important;
    font-size: 0.85rem !important;
}
[data-testid="stTabs"] [data-testid="stTab"][aria-selected="true"] {
    color: #C9A227 !important;
    border-bottom: 2px solid #C9A227 !important;
}

/* ---- Cards & Metrics — versão dourada elevada ---- */
.pa-card, .pa-metric {
    background: linear-gradient(180deg, #FFFFFF 0%, #FBF7EA 100%) !important;
    border: 1px solid #EFE3B8 !important;
    border-top: 3px solid #C9A227 !important;
    border-radius: 14px !important;
    padding: 1.2rem !important;
    margin-bottom: 0.8rem !important;
    box-shadow: 0 2px 6px rgba(140,110,20,0.07), 0 1px 2px rgba(201,162,39,0.10);
    transition: transform 0.18s ease, box-shadow 0.18s ease, border-color 0.18s ease;
}
.pa-card:hover, .pa-metric:hover {
    transform: translateY(-3px);
    box-shadow: 0 12px 26px rgba(140,110,20,0.14), 0 3px 10px rgba(201,162,39,0.22);
    border-color: #C9A227 !important;
}

.pa-metric-value {
    font-size: 24px !important;
    font-weight: 800 !important;
    color: #1A1A1A !important;
    letter-spacing: -0.01em !important;
}
.pa-metric-label {
    font-size: 11px !important;
    font-weight: 700 !important;
    color: #9C8A46 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.07em !important;
    margin-top: 2px !important;
}

.pa-card-header {
    font-size: 0.95rem !important;
    font-weight: 700 !important;
    color: #1A1A1A !important;
    margin-bottom: 4px !important;
}
.pa-card-sub {
    font-size: 0.8rem !important;
    color: #6B6B6B !important;
    font-weight: 500 !important;
}

[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input,
[data-testid="stSelectbox"] > div > div {
    background: #FFFFFF !important;
    border: 1px solid #E8E8E8 !important;
    color: #1A1A1A !important;
    border-radius: 8px !important;
}

.stButton > button {
    background: transparent !important;
    color: #C9A227 !important;
    border: 1.5px solid #C9A227 !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
}
button[kind="primary"], .stButton [kind="primary"] > button {
    background: #C9A227 !important;
    color: #111111 !important;
    border: none !important;
}

/* ---- Tabela de Assistência ---- */
.assist-table {
    width: 100%;
    border-collapse: collapse;
    font-family: 'Inter', sans-serif;
    font-size: 0.82rem;
}
.assist-table th {
    background: #1A1A1A;
    color: #FFFFFF;
    padding: 8px 10px;
    text-align: center;
    font-weight: 600;
    font-size: 0.75rem;
    border: 1px solid #333;
}
.assist-table th.col-mes {
    background: #2A2A2A;
    text-align: left;
}
.assist-table td {
    padding: 7px 10px;
    border: 1px solid #DEDEDE;
    text-align: center;
    background: #FFFFFF;
    color: #1A1A1A;
}
.assist-table td.col-mes {
    text-align: left;
    font-weight: 500;
    background: #FAFAFA;
}
.assist-table tr.row-total td {
    background: #F0F0F0;
    font-weight: 700;
    border-top: 2px solid #AAAAAA;
}
.assist-table .ano-header {
    background: #C9A227;
    color: #111;
    font-size: 1.1rem;
    font-weight: 800;
    text-align: center;
    padding: 6px;
}
</style>
""", unsafe_allow_html=True)


# =============================================================
# 3. CONSTANTES E LISTAS GLOBAIS
# =============================================================
_AUTH_USERS = {"wendley": "Qmerd@10"}

# -------------------------------------------------------------
# COORDENADAS DO CARTÃO S-21 (em mm, origem no canto inferior
# esquerdo da página A4 — padrão do ReportLab)
#
# AJUSTADO com base nas fotos enviadas (cartão do Wendley e o
# consolidado). Os X's e números estavam caindo fora dos
# quadradinhos/colunas certas. Ver observações em cada bloco.
# -------------------------------------------------------------
PDF_Y_OFFSET    = 0.0
PDF_NOME_Y      = 272.0
PDF_NOME_X      = 24.0
PDF_NASCI_Y     = 265.0
PDF_NASCI_X     = 48.0
PDF_BATISM_Y    = 258.0
PDF_BATISM_X    = 48.0
PDF_CARGO_Y     = 252.0

# Checkboxes "Masculino/Feminino" e "Outras ovelhas/Ungido":
# na 1ª rodada o X ainda caiu um pouco à direita do quadradinho
# (Masculino e Outras ovelhas) — recuado mais ~8mm agora.
# Feminino/Ungido já estavam bons, não foram tocados.
PDF_MASC_X      = 135.0   # era 143.0 (ainda pra direita)
PDF_FEM_X       = 165.0   # sem alteração — já estava bom
PDF_OVELHAS_X   = 135.0   # era 143.0 (ainda pra direita)
PDF_UNGIDO_X    = 165.0   # sem alteração — já estava bom

# Checkboxes de cargo (Ancião / Servo / Pioneiro reg. / Pioneiro
# esp. / Missionário): pequeno ajuste para centralizar o X no
# quadradinho — estava nascendo meio pixel à direita, tocando
# o começo do rótulo.
PDF_ANCIAO_X    = 7.0      # era 9.5
PDF_SERVO_X     = 33.5     # era 35.0
PDF_PREG_X      = 63.5     # era 65.0
PDF_PESP_X      = 98.5     # era 100.0
PDF_MISS_X      = 138.5    # era 140.0

# Telefone de emergência: NÃO fica mais dentro da tabela.
# Na foto ele sobrepunha o texto "(Se for pioneiro ou
# missionário em campo)" do cabeçalho da coluna Horas.
# Movido para o espaço em branco à direita da linha de cargos.
PDF_TEL_X       = 150.0
PDF_TEL_Y       = 238.5

_Y_MAP_BASE = {
    "SETEMBRO":  228.5, "OUTUBRO":   220.5, "NOVEMBRO":  212.5, "DEZEMBRO":  204.5,
    "JANEIRO":   196.5, "FEVEREIRO": 188.5, "MARÇO":     180.5, "ABRIL":     172.5,
    "MAIO":      164.5, "JUNHO":     156.5, "JULHO":     148.5, "AGOSTO":    140.5,
}

PDF_TOTAL_Y        = 134.5  # era 131.5 — subiu ~3mm, estava caindo longe demais da linha "Total"

# Colunas da tabela mensal:
# - PARTICIP: o X do "participou no ministério" estava saindo
#   do quadradinho e quase tocando a coluna "Estudos" (visto na
#   linha de Abril e nos meses do consolidado). Trazido ~5.5mm
#   para a esquerda.
# - HORAS: pequeno ajuste para centralizar melhor o número.
# - OBS: um pouco mais à direita para dar respiro em relação à
#   coluna de Horas (na foto "15" e "Pioneiro Auxiliar"/"63
#   relat." ficavam quase colados).
PDF_COL_PARTICIP_X = 48.0   # era 53.5
PDF_COL_ESTUDOS_X  = 80.5
PDF_COL_PIAUX_X    = 97.5
PDF_COL_HORAS_X    = 117.5  # era 116.5
PDF_COL_OBS_X      = 136.0  # era 133.0

_CARGO_X_MAP = {
    "Ancião":               PDF_ANCIAO_X,
    "Servo ministerial":    PDF_SERVO_X,
    "Pioneiro regular":     PDF_PREG_X,
    "Pioneiro especial":    PDF_PESP_X,
    "Missionário em campo": PDF_MISS_X,
}

_CARGOS_LISTA = [
    "Ancião", "Servo ministerial", "Pioneiro regular",
    "Pioneiro especial", "Missionário em campo"
]

_MESES_ORDEM = [
    "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO",
    "JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL",
    "MAIO", "JUNHO", "JULHO", "AGOSTO"
]

# Meses por ano de serviço (Set–Ago)
_MESES_ANO_SERVICO = [
    "Setembro", "Outubro", "Novembro", "Dezembro",
    "Janeiro", "Fevereiro", "Março", "Abril",
    "Maio", "Junho", "Julho", "Agosto"
]

_GENEROS       = ["", "Masculino", "Feminino"]
_CLASSES       = ["", "Outras ovelhas", "Ungido"]
_STATUS_OPCOES = ["Ativo", "Inativo"]

categorias_lista = ["PUBLICADOR", "PIONEIRO AUXILIAR", "PIONEIRO REGULAR"]

meses_referencia_ordem = [
    "SETEMBRO 2024", "OUTUBRO 2024", "NOVEMBRO 2024", "DEZEMBRO 2024",
    "JANEIRO 2025", "FEVEREIRO 2025", "MARÇO 2025", "ABRIL 2025", "MAIO 2025",
    "JUNHO 2025", "JULHO 2025", "AGOSTO 2025",
    "SETEMBRO 2025", "OUTUBRO 2025", "NOVEMBRO 2025", "DEZEMBRO 2025",
    "JANEIRO 2026", "FEVEREIRO 2026", "MARÇO 2026", "ABRIL 2026", "MAIO 2026",
    "JUNHO 2026", "JULHO 2026", "AGOSTO 2026",
]


# =============================================================
# 4. AUTENTICAÇÃO
# =============================================================
def tela_login():
    st.markdown("""
    <style>
    .stApp { background: linear-gradient(180deg, #FAF7EE 0%, #F2ECD6 100%) !important; }
    </style>
    """, unsafe_allow_html=True)

    col_left, col_center, col_right = st.columns([1, 1.2, 1])
    with col_center:
        st.markdown("""
        <div style="background: #FFFFFF; border: 1px solid #EEE3B8; border-radius: 16px;
            margin-top: 12vh; text-align: center; overflow: hidden;
            box-shadow: 0 14px 34px rgba(140,110,20,0.14);">
          <div style="background: #111111; padding: 9px 0; border-bottom: 3px solid #C9A227;">
            <span style="color: #E9CF6B; font-weight: 800; font-size: 0.7rem; letter-spacing: 0.14em;">
                PARQUE ALIANÇA · PORTAL</span>
          </div>
          <div style="padding: 2.2rem 2rem 2.4rem;">
            <div style="display: flex; justify-content: center; margin-bottom: 1.25rem;">
              <div style="background: #C9A227; border-radius: 12px; width: 54px; height: 54px;
                  display: flex; align-items: center; justify-content: center;
                  font-weight: 700; font-size: 20px; color: #111;">PA</div>
            </div>
            <h2 style="color: #1A1A1A !important; font-size: 22px; font-weight: 700; margin-bottom: 6px;">
                Portal de Relatórios</h2>
            <p style="color: #9C8A46 !important; font-size: 13px; margin-bottom: 0.5rem;">
                Congregação Parque Aliança – 72249</p>
          </div>
        </div>
        """, unsafe_allow_html=True)

        with st.container():
            user  = st.text_input("Usuário", placeholder="Digite seu usuário",
                                   label_visibility="collapsed", key="login_user")
            senha = st.text_input("Senha",   placeholder="Digite sua senha",
                                   type="password", label_visibility="collapsed", key="login_pass")
            entrar = st.button("Acessar Portal", use_container_width=True, type="primary")

        if entrar:
            if _AUTH_USERS.get(user.lower().strip()) == senha:
                st.session_state["autenticado"]    = True
                st.session_state["usuario_logado"] = user.strip().title()
                st.rerun()
            else:
                st.error("Usuário ou senha incorretos.")


# =============================================================
# 5. FUNÇÕES UTILITÁRIAS
# =============================================================
def normalizar_texto(texto):
    if not texto:
        return ""
    return "".join(
        c for c in unicodedata.normalize('NFD', str(texto))
        if unicodedata.category(c) != 'Mn'
    ).lower().strip()


def obter_mes_vigente_str():
    meses = ["JANEIRO","FEVEREIRO","MARÇO","ABRIL","MAIO","JUNHO",
             "JULHO","AGOSTO","SETEMBRO","OUTUBRO","NOVEMBRO","DEZEMBRO"]
    hoje = date.today()
    if hoje.day >= 20:
        return f"{meses[hoje.month - 1]} {hoje.year}"
    else:
        if hoje.month == 1:
            return f"DEZEMBRO {hoje.year - 1}"
        return f"{meses[hoje.month - 2]} {hoje.year}"


def cargos_para_lista(cargo_val):
    if not cargo_val:
        return []
    if isinstance(cargo_val, list):
        return [c for c in cargo_val if c]
    return [cargo_val] if cargo_val else []


def ordenar_df_por_mes(df_input):
    def chave_mes(mes_ref):
        partes = str(mes_ref).upper().split()
        nome_mes = partes[0] if partes else ""
        ano = int(partes[1]) if len(partes) > 1 else 0
        idx = _MESES_ORDEM.index(nome_mes) if nome_mes in _MESES_ORDEM else 99
        return (ano, idx)
    df_sorted = df_input.copy()
    df_sorted["_sort_key"] = df_sorted["mes_referencia"].apply(chave_mes)
    df_sorted = df_sorted.sort_values("_sort_key").drop(columns=["_sort_key"])
    return df_sorted


def normalizar_nome_no_banco(nome_recebido, lista_membros):
    entrada_norm = normalizar_texto(nome_recebido)
    if not entrada_norm or len(entrada_norm) < 2:
        return None

    tokens_entrada = set(entrada_norm.split())
    melhor_match, maior_score = None, 0.0

    for nome_oficial in lista_membros:
        oficial_norm = normalizar_texto(nome_oficial)
        tokens_oficial = oficial_norm.split()

        if entrada_norm == oficial_norm:
            return nome_oficial

        if len(tokens_entrada) == 1:
            primeiro = tokens_oficial[0] if tokens_oficial else ""
            segundo  = tokens_oficial[1] if len(tokens_oficial) > 1 else ""
            if entrada_norm in (primeiro, segundo):
                return nome_oficial

        if tokens_entrada and tokens_entrada.issubset(set(tokens_oficial)):
            score = len(tokens_entrada) / max(len(tokens_oficial), 1) + 0.5
            if score > maior_score:
                maior_score, melhor_match = score, nome_oficial
            continue

        primeiro_oficial = tokens_oficial[0] if tokens_oficial else ""
        for tok in tokens_entrada:
            if tok == primeiro_oficial and len(tok) >= 3:
                score = 0.88
                if score > maior_score:
                    maior_score, melhor_match = score, nome_oficial

        score_fuzzy = SequenceMatcher(None, entrada_norm, oficial_norm).ratio()
        if score_fuzzy > maior_score:
            maior_score, melhor_match = score_fuzzy, nome_oficial

    return melhor_match if maior_score >= 0.82 else None


# =============================================================
# 6. BANCO DE DADOS — CONEXÃO E CACHE
# =============================================================
def inicializar_db():
    if "db" not in st.session_state:
        try:
            key_dict = json.loads(st.secrets["textkey"])
            creds = service_account.Credentials.from_service_account_info(key_dict)
            st.session_state.db = firestore.Client(
                credentials=creds, project="wendleydesenvolvimento")
        except Exception:
            return None
    return st.session_state.db


@st.cache_data(ttl=60, show_spinner=False)
def carregar_membros_cached():
    db = inicializar_db()
    if not db:
        return {}
    return {doc.id: doc.to_dict() for doc in db.collection("membros_v2").stream()}


@st.cache_data(ttl=60, show_spinner=False)
def carregar_relatorios_cached():
    db = inicializar_db()
    if not db:
        return []
    docs = db.collection("relatorios_parque_alianca").stream()
    return [
        {"id": doc.id, **doc.to_dict()}
        for doc in docs
        if doc.to_dict().get("status_validacao") != "EXCLUIDO"
    ]


@st.cache_data(ttl=120, show_spinner=False)
def carregar_anuncios_cached():
    db = inicializar_db()
    if not db:
        return []
    try:
        docs = (db.collection("anuncios")
                  .order_by("data_postagem", direction=firestore.Query.DESCENDING)
                  .stream())
        return [{"id": doc.id, **doc.to_dict()} for doc in docs]
    except Exception:
        return []


@st.cache_data(ttl=60, show_spinner=False)
def carregar_assistencia_cached():
    """Carrega todos os registros de assistência às reuniões."""
    db = inicializar_db()
    if not db:
        return []
    try:
        docs = db.collection("assistencia_reunioes").stream()
        return [{"id": doc.id, **doc.to_dict()} for doc in docs]
    except Exception:
        return []


def carregar_membros():
    return carregar_membros_cached()


def carregar_relatorios():
    return carregar_relatorios_cached()


def carregar_anuncios():
    return carregar_anuncios_cached()


def carregar_assistencia():
    return carregar_assistencia_cached()


# =============================================================
# 7. OPERAÇÕES DE ESCRITA NO BANCO
# =============================================================
def atualizar_membro(nome, categoria, novo=False, extra=None):
    db = inicializar_db()
    if db:
        dados = {"categoria": categoria, "nome_oficial": nome}
        if novo:
            dados["mes_inicio"] = obter_mes_vigente_str()
        if extra:
            dados.update({k: v for k, v in extra.items() if v is not None})
        db.collection("membros_v2").document(nome).set(dados, merge=True)
        carregar_membros_cached.clear()


def deletar_relatorio(relatorio_id):
    db = inicializar_db()
    if not db:
        st.error("Sem conexão com o banco.")
        return
    try:
        db.collection("relatorios_parque_alianca").document(relatorio_id).delete()
    except Exception as e:
        st.error(f"Erro ao deletar: {e}")
        return
    carregar_relatorios_cached.clear()
    carregar_membros_cached.clear()
    st.toast("🗑️ Relatório deletado permanentemente!")
    st.rerun()


def deletar_membro(nome):
    db = inicializar_db()
    if not db:
        st.error("Sem conexão com o banco.")
        return
    try:
        db.collection("membros_v2").document(nome).delete()
    except Exception as e:
        st.error(f"Erro ao deletar membro: {e}")
        return
    carregar_membros_cached.clear()
    carregar_relatorios_cached.clear()
    st.toast(f"🗑️ Membro '{nome}' deletado permanentemente!")
    st.rerun()


def salvar_baixa_manual(nome, mes, horas, estudos):
    db = inicializar_db()
    if db:
        novo_doc = {
            "nome": nome, "mes_referencia": mes, "horas": horas,
            "estudos_biblicos": estudos, "timestamp": firestore.SERVER_TIMESTAMP
        }
        db.collection("relatorios_parque_alianca").add(novo_doc)
        carregar_relatorios_cached.clear()
        st.success(f"✅ Relatório de {nome} adicionado!")
        st.rerun()


def salvar_anuncio(dados):
    db = inicializar_db()
    if not db:
        return False
    dados["data_postagem"] = firestore.SERVER_TIMESTAMP
    db.collection("anuncios").add(dados)
    carregar_anuncios_cached.clear()
    return True


def deletar_anuncio(anuncio_id):
    db = inicializar_db()
    if db:
        db.collection("anuncios").document(anuncio_id).delete()
        carregar_anuncios_cached.clear()
        st.toast("✅ Anúncio deletado!")
        st.rerun()


def salvar_assistencia(tipo_reuniao, ano_referencia, mes, qtd_reunioes, total_assistencia):
    """Salva ou atualiza um registro de assistência no Firestore."""
    db = inicializar_db()
    if not db:
        st.error("Sem conexão com o banco.")
        return False
    doc_id = f"{tipo_reuniao}_{ano_referencia}_{mes}".replace(" ", "_").upper()
    media = round(total_assistencia / qtd_reunioes, 1) if qtd_reunioes > 0 else 0
    dados = {
        "tipo_reuniao":      tipo_reuniao,
        "ano_referencia":    ano_referencia,
        "mes":               mes,
        "qtd_reunioes":      qtd_reunioes,
        "total_assistencia": total_assistencia,
        "media_semana":      media,
        "atualizado_em":     firestore.SERVER_TIMESTAMP,
    }
    db.collection("assistencia_reunioes").document(doc_id).set(dados)
    carregar_assistencia_cached.clear()
    return True


# =============================================================
# 8. GERAÇÃO DE PDF (S-21)
# =============================================================
def gerar_pdf_padrao_s21(nome_cabecalho, categoria_label, dados_rows, membro_info=None):
    path_original = os.path.join(os.path.dirname(__file__), "s21.pdf")
    if not os.path.exists(path_original):
        st.error("Arquivo 's21.pdf' não encontrado na pasta do app.")
        return None

    mi = membro_info or {}
    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=A4)

    can.setFont("Helvetica-Bold", 10)
    can.drawString(PDF_NOME_X * mm, PDF_NOME_Y * mm, str(nome_cabecalho).upper())

    data_nasc = str(mi.get("data_nascimento", "")).strip()
    if data_nasc:
        can.setFont("Helvetica", 9)
        can.drawString(PDF_NASCI_X * mm, PDF_NASCI_Y * mm, data_nasc)

    data_bat = str(mi.get("data_batismo", "")).strip()
    if data_bat:
        can.setFont("Helvetica", 9)
        can.drawString(PDF_BATISM_X * mm, PDF_BATISM_Y * mm, data_bat)

    genero = mi.get("genero", "")
    can.setFont("Helvetica-Bold", 10)
    if genero == "Masculino":
        can.drawString(PDF_MASC_X * mm, PDF_NASCI_Y * mm, "X")
    elif genero == "Feminino":
        can.drawString(PDF_FEM_X * mm, PDF_NASCI_Y * mm, "X")

    classe = mi.get("classe", "")
    if classe == "Outras ovelhas":
        can.drawString(PDF_OVELHAS_X * mm, PDF_BATISM_Y * mm, "X")
    elif classe == "Ungido":
        can.drawString(PDF_UNGIDO_X * mm, PDF_BATISM_Y * mm, "X")

    cargos = cargos_para_lista(mi.get("cargo", ""))
    can.setFont("Helvetica-Bold", 10)
    for cargo in cargos:
        if cargo in _CARGO_X_MAP:
            can.drawString(_CARGO_X_MAP[cargo] * mm, PDF_CARGO_Y * mm, "X")

    tel_emerg = str(mi.get("telefone_emergencia", "")).strip()
    if tel_emerg:
        can.setFont("Helvetica-Bold", 8)
        can.drawString(PDF_TEL_X * mm, PDF_TEL_Y * mm, f"Tel: {tel_emerg}"[:32])

    total_horas = 0
    total_estud = 0

    for _, row in dados_rows.iterrows():
        mes_key = str(row.get('mes_referencia', '')).split()[0].upper()
        y_base  = _Y_MAP_BASE.get(mes_key)
        if y_base is None:
            continue
        y_pos = (y_base + PDF_Y_OFFSET) * mm

        horas = int(row.get('horas', 0))
        estud = int(row.get('estudos_biblicos', 0))
        total_horas += horas
        total_estud += estud

        if horas > 0 or estud > 0:
            can.setFont("Helvetica-Bold", 10)
            can.drawCentredString(PDF_COL_PARTICIP_X * mm, y_pos, "X")

        can.setFont("Helvetica-Bold", 10)
        can.drawCentredString(PDF_COL_ESTUDOS_X * mm, y_pos, str(estud))

        categoria_do_mes = str(row.get('cat_oficial', '')).upper()
        if categoria_do_mes == "PIONEIRO AUXILIAR":
            can.drawCentredString(PDF_COL_PIAUX_X * mm, y_pos, "X")

        can.drawCentredString(PDF_COL_HORAS_X * mm, y_pos, str(horas))

        obs_normal = str(row.get('observacoes', ''))
        obs_normal = obs_normal if obs_normal.lower() not in ('nan', '', 'none') else ''

        if categoria_do_mes == "PIONEIRO AUXILIAR":
            obs_final = f"Pion. Auxiliar | {obs_normal}" if obs_normal else "Pioneiro Auxiliar"
        else:
            obs_final = obs_normal

        if obs_final:
            can.setFont("Helvetica", 8)
            can.drawString(PDF_COL_OBS_X * mm, y_pos, obs_final[:32])

        can.setFont("Helvetica-Bold", 10)

    if total_horas > 0:
        can.setFont("Helvetica-Bold", 10)
        can.drawCentredString(PDF_COL_HORAS_X * mm, PDF_TOTAL_Y * mm, str(total_horas))
    if total_estud > 0:
        can.setFont("Helvetica-Bold", 10)
        can.drawCentredString(PDF_COL_ESTUDOS_X * mm, PDF_TOTAL_Y * mm, str(total_estud))

    can.save()
    packet.seek(0)

    reader_original = PdfReader(open(path_original, "rb"))
    writer = PdfWriter()
    pagina_base = reader_original.pages[0]
    pagina_base.merge_page(PdfReader(packet).pages[0])
    writer.add_page(pagina_base)
    output = io.BytesIO()
    writer.write(output)
    return output.getvalue()


def gerar_zip_pendentes(pendentes, mes, membros_db, df_todos):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for nome in pendentes:
            mi = membros_db.get(nome, {})
            df_hist = df_todos[
                (df_todos['nome_oficial'] == nome) &
                (df_todos['status_validacao'] == "IDENTIFICADO")
            ]
            df_hist = ordenar_df_por_mes(df_hist) if not df_hist.empty else pd.DataFrame()
            pdf = gerar_pdf_padrao_s21(nome, mi.get('categoria', 'PUBLICADOR'),
                                       df_hist, membro_info=mi)
            if pdf:
                nome_arq = "".join(c for c in nome if c.isalnum() or c in (' ', '_', '-')
                                   ).strip().replace(' ', '_')
                zf.writestr(f"Pendente_{nome_arq}.pdf", pdf)
    return buf.getvalue()


# =============================================================
# 9. FUNÇÕES DE ANÚNCIOS (HTML)
# =============================================================
def gerar_html_agenda(d):
    C_CANT  = "#1a78b4"
    C_TES   = "#1a3566"
    C_TES_I = "#1a5fa8"
    C_MIN   = "#8a6200"
    C_MIN_I = "#a07800"
    C_NVC   = "#cc0000"
    C_NVC_I = "#1a5fa8"

    def row(num, titulo, duracao, bg, cor_item):
        if not str(titulo).strip():
            return ""
        dur = (f'<br><span style="font-size:12px;color:#888;margin-left:4px;">({duracao})</span>'
               if str(duracao).strip() else "")
        return (f'<div style="padding:6px 14px;background:{bg};border-bottom:1px solid #e8e8e8;">'
                f'<span style="color:{cor_item};font-weight:bold;">{num}. {titulo}</span>{dur}'
                f'</div>')

    def sec_header(texto, bg):
        return (f'<div style="background:{bg};color:white;padding:9px 14px;'
                f'font-weight:bold;font-size:14.5px;letter-spacing:0.3px;">{texto}</div>')

    html = ('<div style="font-family:Arial,Helvetica,sans-serif;max-width:480px;'
            'border:1px solid #ccc;border-radius:10px;overflow:hidden;'
            'box-shadow:0 2px 8px rgba(0,0,0,0.12);margin:auto;">')
    html += f'<div style="padding:12px 14px 8px;background:#ffffff;">'
    html += f'<div style="font-size:19px;font-weight:bold;color:#111;">{d.get("data_texto","")}</div>'
    if d.get("escritura"):
        html += f'<div style="color:{C_CANT};font-size:13px;font-weight:bold;margin-top:2px;">{d["escritura"]}</div>'
    html += '</div>'
    html += '<hr style="margin:0;border:0;border-top:1px solid #ddd;">'

    if d.get("cantico_abertura"):
        html += (f'<div style="padding:7px 14px;font-size:13px;background:#fff;">'
                 f'<span style="color:{C_CANT};font-weight:bold;">Cântico {d["cantico_abertura"]}</span>'
                 f' e oração | <strong>Comentários iniciais</strong> (1 min)</div>')

    html += '<div style="margin-top:8px;">'
    html += sec_header("TESOUROS DA PALAVRA DE DEUS", C_TES)
    for it in d.get("tesouros", []):
        html += row(it["num"], it["titulo"], it.get("duracao", ""), "#f0f4ff", C_TES_I)
    html += '</div>'

    html += '<div style="margin-top:8px;">'
    html += sec_header("FAÇA SEU MELHOR NO MINISTÉRIO", C_MIN)
    for it in d.get("ministerio", []):
        html += row(it["num"], it["titulo"], it.get("duracao", ""), "#fffcf0", C_MIN_I)
    html += '</div>'

    html += '<div style="margin-top:8px;">'
    html += sec_header("NOSSA VIDA CRISTÃ", C_NVC)
    if d.get("cantico_meio"):
        html += (f'<div style="padding:6px 14px;background:#fff5f5;border-bottom:1px solid #e8e8e8;">'
                 f'<span style="color:{C_CANT};font-weight:bold;">Cântico {d["cantico_meio"]}</span></div>')
    for it in d.get("vida_crista", []):
        html += row(it["num"], it["titulo"], it.get("duracao", ""), "#fff5f5", C_NVC_I)
    html += '</div>'

    if d.get("cantico_final"):
        html += (f'<hr style="margin:0;border:0;border-top:1px solid #ddd;">'
                 f'<div style="padding:9px 14px;font-size:13px;background:#fff;">'
                 f'<strong>Comentários finais</strong> (3 min) | '
                 f'<span style="color:{C_CANT};font-weight:bold;">Cântico {d["cantico_final"]}</span>'
                 f' e oração</div>')

    html += '</div>'
    return html


# =============================================================
# 10. PROCESSAMENTO DE DADOS (RELATÓRIOS)
# =============================================================
def processar_dataframe(relatorios_brutos, membros_db):
    if not relatorios_brutos:
        return pd.DataFrame()

    df = pd.DataFrame(relatorios_brutos)

    if 'status_validacao' in df.columns:
        df = df[df['status_validacao'] != "EXCLUIDO"].copy()

    if df.empty:
        return pd.DataFrame()

    df['horas']            = pd.to_numeric(df.get('horas', 0), errors='coerce').fillna(0)
    df['estudos_biblicos'] = pd.to_numeric(df.get('estudos_biblicos', 0), errors='coerce').fillna(0)
    df['mes_referencia']   = df['mes_referencia'].str.upper()

    lista_nomes = list(membros_db.keys())

    def validar_envio(row):
        nome_oficial = normalizar_nome_no_banco(row['nome'], lista_nomes)
        if nome_oficial:
            dados_m = membros_db[nome_oficial]
            cat_mes = row.get('categoria_mes')
            if pd.notna(cat_mes) and cat_mes in categorias_lista:
                cat_final = cat_mes
            else:
                cat_final = dados_m.get('categoria', 'PUBLICADOR')
                if cat_final not in categorias_lista:
                    cat_final = 'PUBLICADOR'
            return pd.Series([nome_oficial, cat_final, "IDENTIFICADO"])
        return pd.Series([row['nome'], "DESCONHECIDO", "TRIAGEM"])

    df[['nome_oficial', 'cat_oficial', 'status_validacao']] = df.apply(validar_envio, axis=1)
    return df


# =============================================================
# TAB: REGISTRO DE ASSISTÊNCIA (S-88-T)
# Reformulação completa — layout fiel ao formulário original
# =============================================================
# Dependências: streamlit, google-cloud-firestore, openpyxl
# Integração: chame render_tab_assistencia(db, congregacao_id)
#             de dentro do seu bloco de abas principal.
# =============================================================

import io
import streamlit as st
from datetime import datetime

MESES_ORDEM = [
    "Setembro", "Outubro", "Novembro", "Dezembro",
    "Janeiro", "Fevereiro", "Março", "Abril",
    "Maio", "Junho", "Julho", "Agosto",
]

TIPOS = ["Reunião do Meio de Semana", "Reunião do Fim de Semana"]

# ─────────────────────────────────────────────────────────────
# FIRESTORE helpers
# ─────────────────────────────────────────────────────────────

def _col(db, cong_id: str):
    return db.collection("congregacoes").document(cong_id).collection("assistencia")


def _doc_id(tipo: str, ano_ref: str) -> str:
    slug = tipo.lower().replace(" ", "_").replace("ã", "a").replace("é", "e")
    return f"{slug}_{ano_ref.replace('/', '-')}"


def _carregar(db, cong_id: str, tipo: str, ano_ref: str) -> dict:
    """Retorna dict {mes: {qtd, total}} ou {} se não existir."""
    try:
        doc = _col(db, cong_id).document(_doc_id(tipo, ano_ref)).get()
        if doc.exists:
            data = doc.to_dict() or {}
            return data.get("meses", {})
    except Exception:
        pass
    return {}


def _salvar(db, cong_id: str, tipo: str, ano_ref: str, meses: dict) -> bool:
    """Persiste o documento no Firestore. Retorna True em caso de sucesso."""
    try:
        _col(db, cong_id).document(_doc_id(tipo, ano_ref)).set({
            "tipo_reuniao": tipo,
            "ano_referencia": ano_ref,
            "meses": meses,
            "atualizado_em": datetime.utcnow().isoformat(),
        })
        return True
    except Exception as exc:
        st.error(f"Erro ao salvar: {exc}")
        return False


# ─────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────

def _gerar_excel(tipo: str, ano_ref: str, meses_data: dict) -> bytes | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Assistência"

        borda = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        cinza   = PatternFill("solid", fgColor="DDDDDD")
        dourado = PatternFill("solid", fgColor="C9A84C")
        azul    = PatternFill("solid", fgColor="2C3E50")

        # Título principal
        ws.merge_cells("A1:D1")
        ws["A1"] = f"REGISTRO DA ASSISTÊNCIA ÀS REUNIÕES CONGREGACIONAIS"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = azul
        ws["A1"].alignment = Alignment(horizontal="center")

        # Sub-título
        ws.merge_cells("A2:D2")
        ws["A2"] = f"{tipo.upper()}   —   ANO DE SERVIÇO: {ano_ref}"
        ws["A2"].font = Font(bold=True, size=11, color="C9A84C")
        ws["A2"].fill = azul
        ws["A2"].alignment = Alignment(horizontal="center")

        # Cabeçalhos
        headers = ["Mês", "Qtd. Reuniões", "Assistência Total", "Média por Semana"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(bold=True, size=10)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.fill = cinza
            c.border = borda

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.row_dimensions[3].height = 30

        soma_qtd = soma_tot = meses_com_dado = 0

        for i, mes in enumerate(MESES_ORDEM, 4):
            dados = meses_data.get(mes, {})
            qtd   = dados.get("qtd", 0) or 0
            total = dados.get("total", 0) or 0
            media = round(total / qtd, 1) if qtd > 0 else 0

            valores = [mes, qtd or "", total or "", media or ""]
            for col, v in enumerate(valores, 1):
                c = ws.cell(row=i, column=col, value=v)
                c.alignment = Alignment(
                    horizontal="left" if col == 1 else "center"
                )
                c.border = borda

            soma_qtd += qtd
            soma_tot += total
            if qtd > 0:
                meses_com_dado += 1

        # Linha de totais
        media_geral = round(soma_tot / meses_com_dado, 1) if meses_com_dado > 0 else 0
        row_tot = 4 + len(MESES_ORDEM)
        totais = ["Assistência média por mês", soma_qtd, soma_tot, media_geral]
        for col, v in enumerate(totais, 1):
            c = ws.cell(row=row_tot, column=col, value=v)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = dourado
            c.alignment = Alignment(
                horizontal="left" if col == 1 else "center"
            )
            c.border = borda

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        return None


# ─────────────────────────────────────────────────────────────
# CSS injetado uma única vez
# ─────────────────────────────────────────────────────────────

def _inject_css():
    st.markdown("""
<style>
/* ── cabeçalho do formulário S-88 ── */
.s88-header {
    background: #2c3e50;
    color: #ffffff;
    padding: 10px 18px 6px;
    border-radius: 8px 8px 0 0;
    margin-bottom: 0;
}
.s88-header h3 {
    margin: 0 0 2px;
    font-size: 1.05rem;
    font-weight: 700;
    letter-spacing: .04em;
    color: #C9A84C;
}
.s88-header p  {
    margin: 0;
    font-size: .78rem;
    color: #b0bec5;
}

/* ── tabela do formulário ── */
.s88-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 0.82rem;
}
.s88-table th {
    background: #2c3e50;
    color: #C9A84C;
    text-align: center;
    padding: 7px 4px;
    font-size: 0.78rem;
    font-weight: 600;
    border: 1px solid #405060;
}
.s88-table td {
    border: 1px solid #ddd;
    padding: 3px 6px;
    vertical-align: middle;
    text-align: center;
}
.s88-table td.mes-label {
    text-align: left;
    font-weight: 500;
    color: #2c3e50;
    padding-left: 10px;
    white-space: nowrap;
}
.s88-table tr.totais-row td {
    background: #C9A84C;
    color: #fff;
    font-weight: 700;
    font-size: 0.8rem;
}

/* compactar inputs numéricos dentro da tabela */
div[data-testid="stNumberInput"] {
    margin: 0 !important;
}
div[data-testid="stNumberInput"] input {
    padding: 3px 6px !important;
    font-size: 0.82rem !important;
    text-align: center !important;
}
div[data-testid="stNumberInput"] > div > div:last-child {
    display: none !important;   /* esconde botões +/- (opcionais) */
}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# Bloco de um tipo de reunião (meio / fim de semana)
# ─────────────────────────────────────────────────────────────

def _bloco_reuniao(db, cong_id: str, tipo: str, ano_ref: str, prefixo: str):
    """
    Renderiza o cabeçalho + tabela editável para um tipo de reunião.
    prefixo: string curta única para key dos widgets (ex: "msem" / "fsem")
    Retorna (meses_dict, soma_qtd, soma_tot, media_geral).
    """
    # Carrega dados do Firestore (uma vez por sessão)
    cache_key = f"assis_{prefixo}_{ano_ref}"
    if cache_key not in st.session_state:
        st.session_state[cache_key] = _carregar(db, cong_id, tipo, ano_ref)

    meses_data: dict = st.session_state[cache_key]

    # ── Cabeçalho visual ──
    st.markdown(f"""
<div class="s88-header">
  <h3>{tipo.upper()}</h3>
  <p>Ano de serviço: <strong style="color:#e0e0e0">{ano_ref}</strong></p>
</div>
""", unsafe_allow_html=True)

    # ── Cabeçalho da tabela (HTML puro) ──
    st.markdown("""
<table class="s88-table">
  <thead>
    <tr>
      <th style="width:22%">Mês</th>
      <th style="width:26%">Qtd. de Reuniões</th>
      <th style="width:26%">Assistência Total</th>
      <th style="width:26%">Média por Semana</th>
    </tr>
  </thead>
</table>
""", unsafe_allow_html=True)

    # ── Linhas editáveis ──
    soma_qtd = soma_tot = meses_com_dado = 0
    medias = []

    for mes in MESES_ORDEM:
        dados_mes  = meses_data.get(mes, {})
        val_qtd    = int(dados_mes.get("qtd",   0) or 0)
        val_total  = int(dados_mes.get("total", 0) or 0)

        col_mes, col_qtd, col_tot, col_med = st.columns([2.2, 2.6, 2.6, 2.6])

        with col_mes:
            st.markdown(f"""
<div style="
    padding: 6px 10px;
    font-weight: 500;
    font-size: 0.83rem;
    color: var(--text-color);
    border-bottom: 1px solid #e0e0e0;
    height: 44px;
    display: flex;
    align-items: center;
">{mes}</div>""", unsafe_allow_html=True)

        with col_qtd:
            qtd = st.number_input(
                "qtd", min_value=0, max_value=9999,
                value=val_qtd, step=1,
                label_visibility="collapsed",
                key=f"{prefixo}_{mes}_qtd",
            )

        with col_tot:
            total = st.number_input(
                "tot", min_value=0, max_value=99999,
                value=val_total, step=1,
                label_visibility="collapsed",
                key=f"{prefixo}_{mes}_tot",
            )

        # calcula média
        media = round(total / qtd, 1) if qtd > 0 else 0.0
        with col_med:
            st.markdown(f"""
<div style="
    padding: 6px 10px;
    font-size: 0.83rem;
    color: {'#2c3e50' if media > 0 else '#aaa'};
    border-bottom: 1px solid #e0e0e0;
    height: 44px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-weight: {'600' if media > 0 else '400'};
">{media if media > 0 else '—'}</div>""", unsafe_allow_html=True)

        # acumula
        meses_data[mes] = {"qtd": int(qtd), "total": int(total)}
        soma_qtd += int(qtd)
        soma_tot += int(total)
        if int(qtd) > 0:
            meses_com_dado += 1
        medias.append(media)

    # Atualiza estado
    st.session_state[cache_key] = meses_data

    # ── Linha de totais ──
    media_geral = round(soma_tot / meses_com_dado, 1) if meses_com_dado > 0 else 0.0
    st.markdown(f"""
<table class="s88-table">
  <tbody>
    <tr class="totais-row">
      <td style="width:22%;text-align:left;padding-left:10px">
        Assistência média por mês
      </td>
      <td style="width:26%">{soma_qtd if soma_qtd else '—'}</td>
      <td style="width:26%">{soma_tot if soma_tot else '—'}</td>
      <td style="width:26%">{media_geral if media_geral else '—'}</td>
    </tr>
  </tbody>
</table>
""", unsafe_allow_html=True)

    return meses_data, soma_qtd, soma_tot, media_geral


# ─────────────────────────────────────────────────────────────
# Função principal — chame do seu arquivo de abas
# ─────────────────────────────────────────────────────────────

def render_tab_assistencia(db, congregacao_id: str):
    """
    Ponto de entrada da aba.
    Exemplo de uso:
        with tab_assistencia:
            render_tab_assistencia(db, CONGREGACAO_ID)
    """
    _inject_css()

    # ── Título da aba ──
    st.markdown("""
<div style="
    background:#2c3e50;
    color:#C9A84C;
    padding:14px 20px;
    border-radius:8px;
    margin-bottom:20px;
">
  <h2 style="margin:0;font-size:1.15rem;letter-spacing:.05em">
    📋 REGISTRO DA ASSISTÊNCIA ÀS REUNIÕES CONGREGACIONAIS
  </h2>
  <p style="margin:4px 0 0;font-size:.78rem;color:#90a4ae">
    Formulário S-88-T · Preencha mês a mês e salve no Firestore
  </p>
</div>
""", unsafe_allow_html=True)

    # ── Seleção do Ano de Serviço ──
    ano_atual  = datetime.now().year
    anos_opcao = [f"{a}/{a+1}" for a in range(ano_atual - 5, ano_atual + 2)]
    # detecta ano de serviço atual (setembro começa o ano JW)
    mes_atual  = datetime.now().month
    ano_padrao = f"{ano_atual}/{ano_atual+1}" if mes_atual >= 9 else f"{ano_atual-1}/{ano_atual}"

    col_sel, col_spacer = st.columns([2, 5])
    with col_sel:
        ano_ref = st.selectbox(
            "Ano de Serviço",
            options=anos_opcao,
            index=anos_opcao.index(ano_padrao) if ano_padrao in anos_opcao else 0,
            key="s88_ano_ref",
        )

    st.markdown("---")

    # ── Dois blocos lado a lado (ou empilhados em telas pequenas) ──
    col_msem, col_gap, col_fsem = st.columns([1, 0.04, 1])

    with col_msem:
        dados_msem, qtd_msem, tot_msem, med_msem = _bloco_reuniao(
            db, congregacao_id,
            tipo="Reunião do Meio de Semana",
            ano_ref=ano_ref,
            prefixo="msem",
        )

    with col_gap:
        st.markdown(
            "<div style='border-left:2px solid #ddd;height:100%;min-height:600px'></div>",
            unsafe_allow_html=True,
        )

    with col_fsem:
        dados_fsem, qtd_fsem, tot_fsem, med_fsem = _bloco_reuniao(
            db, congregacao_id,
            tipo="Reunião do Fim de Semana",
            ano_ref=ano_ref,
            prefixo="fsem",
        )

    st.markdown("---")

    # ── Botões de ação ──
    col_salvar, col_excel_m, col_excel_f, col_imprimir = st.columns([1.5, 1.5, 1.5, 1])

    with col_salvar:
        if st.button("💾 Salvar no Firestore", type="primary", use_container_width=True):
            ok_m = _salvar(db, congregacao_id, "Reunião do Meio de Semana", ano_ref, dados_msem)
            ok_f = _salvar(db, congregacao_id, "Reunião do Fim de Semana",  ano_ref, dados_fsem)
            if ok_m and ok_f:
                st.success("✅ Dados salvos com sucesso!")
            else:
                st.error("Falha ao salvar um ou ambos os registros.")

    with col_excel_m:
        bytes_m = _gerar_excel("Reunião do Meio de Semana", ano_ref, dados_msem)
        if bytes_m:
            st.download_button(
                label="📥 Excel — Meio de Semana",
                data=bytes_m,
                file_name=f"assistencia_meio_semana_{ano_ref.replace('/','-')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.info("openpyxl não instalado")

    with col_excel_f:
        bytes_f = _gerar_excel("Reunião do Fim de Semana", ano_ref, dados_fsem)
        if bytes_f:
            st.download_button(
                label="📥 Excel — Fim de Semana",
                data=bytes_f,
                file_name=f"assistencia_fim_semana_{ano_ref.replace('/','-')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.info("openpyxl não instalado")

    with col_imprimir:
        # Impressão via JS — abre janela de impressão do navegador
        st.markdown("""
<button onclick="window.print()" style="
    width:100%;
    padding:8px 0;
    background:#2c3e50;
    color:#C9A84C;
    border:none;
    border-radius:6px;
    font-size:0.85rem;
    font-weight:600;
    cursor:pointer;
    letter-spacing:.03em;
">🖨️ Imprimir</button>
""", unsafe_allow_html=True)

    # ── Resumo visual compacto (opcional) ──
    with st.expander("📊 Resumo do ano de serviço", expanded=False):
        r1, r2, r3, r4 = st.columns(4)
        r1.metric("Total reuniões (meio)",    qtd_msem)
        r2.metric("Assistência total (meio)", tot_msem)
        r3.metric("Total reuniões (fim)",     qtd_fsem)
        r4.metric("Assistência total (fim)",  tot_fsem)

        r5, r6 = st.columns(2)
        r5.metric("Média geral — meio de semana", f"{med_msem}")
        r6.metric("Média geral — fim de semana",  f"{med_fsem}")


# =============================================================
# 12. SIDEBAR
# =============================================================
def renderizar_sidebar(df, mes_vigente):
    with st.sidebar:
        st.markdown(f"""
        <div class="sidebar-brand">
            <div style="font-size:2rem;margin-bottom:4px;">🕊️</div>
            <div class="sidebar-brand-title">Parque Aliança</div>
            <div class="sidebar-brand-sub">Gestão · v5.2</div>
        </div>
        <hr class="sidebar-divider">
        """, unsafe_allow_html=True)

        st.markdown(
            '<p style="color:#6b7280;font-size:0.7rem;text-transform:uppercase;'
            'letter-spacing:0.1em;font-weight:700;margin-bottom:4px;">Mês de Análise</p>',
            unsafe_allow_html=True
        )

        meses_disponiveis = sorted(df['mes_referencia'].unique()) if not df.empty else [mes_vigente]
        idx_default = len(meses_disponiveis) - 1
        if mes_vigente in meses_disponiveis:
            idx_default = meses_disponiveis.index(mes_vigente)

        mes_sel = st.selectbox(
            "Mês", meses_disponiveis, index=idx_default, label_visibility="collapsed"
        )

        eh_vigente = (mes_sel == mes_vigente)
        if eh_vigente:
            st.markdown("""
            <div class="mes-badge">
                <span class="mes-dot"></span>MÊS VIGENTE
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown("""
            <div style="display:inline-flex;align-items:center;gap:6px;
                background:#1f2937;border:1px solid #374151;border-radius:999px;
                padding:5px 14px;font-size:0.75rem;font-weight:700;color:#6b7280;">
                📅 HISTÓRICO</div>""", unsafe_allow_html=True)

        st.markdown('<hr class="sidebar-divider">', unsafe_allow_html=True)

        if not df.empty:
            df_mes_side = df[df['mes_referencia'] == mes_sel]
            df_id_side  = df_mes_side[df_mes_side['status_validacao'] == "IDENTIFICADO"]
            df_tri_side = df_mes_side[df_mes_side['status_validacao'] == "TRIAGEM"]

            st.markdown(f"""
            <div style="display:grid;gap:6px;">
              <div class="pa-metric">
                <div class="pa-metric-value">{len(df_id_side)}</div>
                <div class="pa-metric-label">Identificados</div>
              </div>
              <div class="pa-metric">
                <div class="pa-metric-value" style="color:#ef4444">{len(df_tri_side)}</div>
                <div class="pa-metric-label">Em triagem</div>
              </div>
            </div>""", unsafe_allow_html=True)

        st.markdown('<hr class="sidebar-divider">', unsafe_allow_html=True)

        usuario = st.session_state.get("usuario_logado", "Admin")
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:10px;padding:6px 0;">
          <div style="width:32px;height:32px;border-radius:50%;
            background:linear-gradient(135deg,#d97706,#f59e0b);
            display:flex;align-items:center;justify-content:center;
            font-weight:800;font-size:0.85rem;color:#000;">{usuario[0].upper()}</div>
          <div>
            <div style="font-weight:700;font-size:0.82rem;color:#f9fafb;">{usuario}</div>
            <div style="font-size:0.68rem;color:#6b7280;text-transform:uppercase;
                letter-spacing:0.05em;">Administrador</div>
          </div>
        </div>""", unsafe_allow_html=True)

        if st.button("Sair", use_container_width=True):
            for k in ["autenticado", "usuario_logado"]:
                st.session_state.pop(k, None)
            st.rerun()

    return mes_sel


# =============================================================
# 13. ABA: RELATÓRIOS
# =============================================================
def aba_relatorios(df_ok, df_mes, mes_sel, membros_db, df):
    st.markdown(f"### 📋 Relatórios de {mes_sel}")
    sub_rel = st.tabs(["👤 PUBLICADOR", "🌟 P. AUXILIAR", "💎 P. REGULAR", "⏳ PENDÊNCIAS"])

    entregaram = set(df_ok['nome_oficial'].unique()) if not df_ok.empty else set()

    for i, cat in enumerate(categorias_lista):
        with sub_rel[i]:
            df_cat = df_ok[df_ok['cat_oficial'] == cat] if not df_ok.empty else pd.DataFrame()

            if df_cat.empty:
                st.info(f"Nenhum envio de {cat} em {mes_sel}.")
            else:
                c1, c2, c3 = st.columns(3)
                with c1:
                    st.markdown(f"""
                    <div class="pa-metric">
                      <div class="pa-metric-value">{len(df_cat)}</div>
                      <div class="pa-metric-label">Relatórios</div>
                    </div>""", unsafe_allow_html=True)
                with c2:
                    st.markdown(f"""
                    <div class="pa-metric">
                      <div class="pa-metric-value">{int(df_cat['horas'].sum())}h</div>
                      <div class="pa-metric-label">Total de Horas</div>
                    </div>""", unsafe_allow_html=True)
                with c3:
                    st.markdown(f"""
                    <div class="pa-metric">
                      <div class="pa-metric-value">{int(df_cat['estudos_biblicos'].sum())}</div>
                      <div class="pa-metric-label">Estudos Bíblicos</div>
                    </div>""", unsafe_allow_html=True)

                st.markdown("")
                df_cat_sorted = df_cat.sort_values('nome_oficial')
                cols = st.columns(4)
                for idx, (_, r) in enumerate(df_cat_sorted.iterrows()):
                    with cols[idx % 4]:
                        st.markdown(
                            f'<div class="pa-card">'
                            f'<div class="pa-card-header">{r["nome_oficial"]}</div>'
                            f'<div class="pa-card-sub">⏱ {int(r["horas"])}h &nbsp;·&nbsp; 📚 {int(r["estudos_biblicos"])}</div>'
                            f'</div>',
                            unsafe_allow_html=True)

    with sub_rel[3]:
        idx_mes_sel = (meses_referencia_ordem.index(mes_sel)
                       if mes_sel in meses_referencia_ordem else 99)

        for cat in categorias_lista:
            pendentes = []
            for n, d_m in membros_db.items():
                inicio  = d_m.get('mes_inicio', 'SETEMBRO 2025')
                idx_ini = (meses_referencia_ordem.index(inicio)
                           if inicio in meses_referencia_ordem else 0)
                if (d_m.get('categoria') == cat
                        and n not in entregaram
                        and idx_mes_sel >= idx_ini
                        and d_m.get('status', 'Ativo') == 'Ativo'):
                    pendentes.append(n)

            pendentes = sorted(pendentes)
            if not pendentes:
                continue

            icone = '👤' if cat == 'PUBLICADOR' else ('💎' if 'AUXILIAR' in cat else '⭐')
            with st.expander(f"{icone} {cat} — {len(pendentes)} pendente(s)", expanded=False):
                col_btn_baixa, _ = st.columns([2, 3])
                with col_btn_baixa:
                    if st.button(f"✔ Dar Baixa em Todos ({len(pendentes)})",
                                 key=f"baixa_all_{cat}_{mes_sel}", type="primary"):
                        db = inicializar_db()
                        if db:
                            batch = db.batch()
                            for p in pendentes:
                                doc_ref = db.collection("relatorios_parque_alianca").document()
                                batch.set(doc_ref, {
                                    "nome": p, "mes_referencia": mes_sel,
                                    "horas": 0, "estudos_biblicos": 0,
                                    "timestamp": firestore.SERVER_TIMESTAMP
                                })
                            batch.commit()
                            carregar_relatorios_cached.clear()
                            st.success(f"✅ Baixa realizada para {len(pendentes)} publicadores(as)!")
                            st.rerun()

                st.markdown("---")
                for p in pendentes:
                    c1, c2, c3, c4 = st.columns([3, 1, 1, 2])
                    c1.markdown(f"**{p}**")
                    h_manual = c2.number_input("H", min_value=0, step=1,
                                               key=f"h_man_{p}_{mes_sel}")
                    e_manual = c3.number_input("E", min_value=0, step=1,
                                               key=f"e_man_{p}_{mes_sel}")
                    if c4.button("✔ Dar Baixa", key=f"btn_man_{p}_{mes_sel}"):
                        salvar_baixa_manual(p, mes_sel, h_manual, e_manual)


# =============================================================
# 14. ABA: TRIAGEM
# =============================================================
def aba_triagem(df_mes, membros_db):
    df_triagem = (df_mes[df_mes['status_validacao'] == "TRIAGEM"]
                  if not df_mes.empty else pd.DataFrame())

    if df_triagem.empty:
        st.markdown("""
        <div style="text-align:center;padding:3rem 1rem;">
          <div style="font-size:3rem;margin-bottom:0.5rem;">✅</div>
          <div style="font-size:1.1rem;font-weight:700;color:#6ee7b7;">Tudo limpo!</div>
          <div style="color:#6b7280;font-size:0.85rem;margin-top:4px;">
              Nenhum relatório em triagem para este mês.</div>
        </div>""", unsafe_allow_html=True)
        return

    st.markdown(f"""
    <div style="margin-bottom:1.5rem;">
      <div style="font-size:0.75rem;font-weight:700;color:#f59e0b;
          text-transform:uppercase;letter-spacing:0.1em;margin-bottom:4px;">
          ⚠️ Triagem — {len(df_triagem)} item(s)
      </div>
      <div style="color:#6b7280;font-size:0.82rem;">
          Estes relatórios precisam de validação manual.
      </div>
    </div>""", unsafe_allow_html=True)

    nomes_db = sorted(list(membros_db.keys()))

    for _, row in df_triagem.iterrows():
        sugestao = normalizar_nome_no_banco(row['nome'], nomes_db)
        idx_sug  = nomes_db.index(sugestao) + 1 if sugestao else 0
        conf_str = "Auto-sugerido" if sugestao else "Não reconhecido"

        with st.container(border=True):
            col_info, col_badge = st.columns([4, 1])
            with col_info:
                st.markdown(f"""
                <div style="margin-bottom:8px;">
                  <span style="font-weight:700;color:#f9fafb;font-size:0.95rem;">
                      "{row['nome']}"</span>
                  <span style="color:#6b7280;font-size:0.8rem;margin-left:8px;">
                      · {int(row['horas'])}h · {int(row.get('estudos_biblicos',0))} estudos</span>
                </div>""", unsafe_allow_html=True)
            with col_badge:
                st.markdown(
                    f'<span style="font-size:0.75rem;font-weight:700;color:#f59e0b;">'
                    f'{conf_str}</span>', unsafe_allow_html=True)

            if sugestao:
                st.markdown(f"""
                <div style="background:#1a1200;border:1px solid #92400e;border-radius:8px;
                    padding:6px 12px;margin-bottom:10px;font-size:0.8rem;color:#fbbf24;">
                    💡 Sugestão: <strong>{sugestao}</strong>
                </div>""", unsafe_allow_html=True)

            c1, c2 = st.columns(2)
            vincular = c1.selectbox(
                "Vincular a:", ["-- Novo Membro --"] + nomes_db,
                index=idx_sug, key=f"v_{row['id']}"
            )
            cat_v = c2.selectbox("Categoria:", categorias_lista, key=f"c_{row['id']}")

            col_confirm, col_del = st.columns([2, 1])
            with col_confirm:
                if st.button("✔ Confirmar Vinculação", key=f"b_{row['id']}",
                             type="primary", use_container_width=True):
                    nome_final = row['nome'] if vincular == "-- Novo Membro --" else vincular
                    atualizar_membro(nome_final, cat_v, novo=(vincular == "-- Novo Membro --"))
                    inicializar_db().collection("relatorios_parque_alianca") \
                        .document(row['id']).update({"nome": nome_final})
                    carregar_relatorios_cached.clear()
                    st.rerun()
            with col_del:
                if st.button("🗑 Deletar", key=f"del_{row['id']}", use_container_width=True):
                    deletar_relatorio(row['id'])


# =============================================================
# 15. ABA: CONSOLIDADO
# =============================================================
def aba_consolidado(df, membros_db, mes_vigente, registros_assistencia):
    c1_tab, c2_tab, c3_tab = st.tabs([
        "👤 INDIVIDUAL (HISTÓRICO)",
        "📊 POR CATEGORIA",
        "🏛️ REGISTRO DE ASSISTÊNCIA",
    ])

    # ---- Sub-aba 1: Individual ----
    with c1_tab:
        membros_ord = sorted(list(membros_db.keys()))
        publicador  = st.selectbox("Publicador", membros_ord)

        st.markdown("---")
        st.markdown("#### 📦 Exportar Todos os Cartões S-21")
        st.caption("Gera um ZIP com o cartão histórico completo de **todos** os membros.")

        if st.button("⚙️ Preparar ZIP — Todos os Cartões", use_container_width=True):
            if df.empty:
                st.warning("Nenhum relatório encontrado.")
                st.session_state.pop("zip_todos_cartoes", None)
            else:
                prog = st.progress(0, text="Iniciando...")
                membros_lista = sorted(membros_db.keys())
                buf_all = io.BytesIO()
                count_ok = 0
                total_m  = len(membros_lista)

                with zipfile.ZipFile(buf_all, "w", compression=zipfile.ZIP_DEFLATED) as zf_all:
                    for i, nome_m in enumerate(membros_lista):
                        prog.progress((i + 1) / total_m, text=f"{nome_m} ({i+1}/{total_m})")
                        df_hist_m = df[
                            (df['nome_oficial'] == nome_m) &
                            (df['status_validacao'] == "IDENTIFICADO")
                        ]
                        if df_hist_m.empty:
                            continue
                        df_hist_m = ordenar_df_por_mes(df_hist_m)
                        mi_m  = membros_db.get(nome_m, {})
                        pdf_m = gerar_pdf_padrao_s21(
                            nome_m, mi_m.get('categoria', 'PUBLICADOR'),
                            df_hist_m, membro_info=mi_m
                        )
                        if pdf_m:
                            nome_arq = "".join(
                                c for c in nome_m if c.isalnum() or c in (' ', '_', '-')
                            ).strip().replace(' ', '_')
                            zf_all.writestr(f"S21_{nome_arq}.pdf", pdf_m)
                            count_ok += 1

                prog.empty()
                if count_ok:
                    st.session_state["zip_todos_cartoes"] = buf_all.getvalue()
                    st.session_state["zip_todos_nome"]    = f"S21_Todos_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
                    st.session_state["zip_todos_count"]   = count_ok
                    st.success(f"✅ {count_ok} cartões prontos!")
                else:
                    st.warning("Nenhum PDF gerado.")
                    st.session_state.pop("zip_todos_cartoes", None)

        if "zip_todos_cartoes" in st.session_state:
            st.download_button(
                f"📥 Baixar ZIP ({st.session_state.get('zip_todos_count','?')} cartões)",
                data=st.session_state["zip_todos_cartoes"],
                file_name=st.session_state.get("zip_todos_nome", "S21_Todos.zip"),
                mime="application/zip",
                use_container_width=True,
                type="primary",
            )

        st.markdown("---")
        st.markdown("#### 👤 Cartão Individual")

        if publicador:
            df_hist = df[
                (df['nome_oficial'] == publicador) &
                (df['status_validacao'] == "IDENTIFICADO")
            ] if not df.empty else pd.DataFrame()

            if not df_hist.empty:
                df_hist = ordenar_df_por_mes(df_hist)
                st.dataframe(
                    df_hist[['mes_referencia', 'horas', 'estudos_biblicos']].rename(columns={
                        'mes_referencia': 'Mês', 'horas': 'Horas', 'estudos_biblicos': 'Estudos'
                    }),
                    use_container_width=True, hide_index=True,
                )
                pdf = gerar_pdf_padrao_s21(
                    publicador, membros_db[publicador].get('categoria'),
                    df_hist, membro_info=membros_db[publicador]
                )
                if pdf:
                    st.download_button(
                        "📥 Baixar Cartão S-21", pdf, f"S21_{publicador}.pdf",
                        use_container_width=True,
                    )
            else:
                st.info("Nenhum relatório identificado para este publicador.")

    # ---- Sub-aba 2: Por Categoria ----
    with c2_tab:
        cat_sel = st.selectbox("Categoria", categorias_lista)
        df_cons = df[
            (df['status_validacao'] == "IDENTIFICADO") &
            (df['cat_oficial'] == cat_sel)
        ] if not df.empty else pd.DataFrame()

        if not df_cons.empty:
            resumo = df_cons.groupby('mes_referencia').agg(
                total_relatorios=('id',              'count'),
                total_horas     =('horas',           'sum'),
                total_estudos   =('estudos_biblicos', 'sum'),
            ).reset_index()

            resumo_ord = ordenar_df_por_mes(resumo)

            def obs_col(row):
                if row['mes_referencia'] == mes_vigente:
                    return f"📌 {int(row['total_relatorios'])} relatórios entregues"
                return ""
            resumo_ord['observacao'] = resumo_ord.apply(obs_col, axis=1)

            st.dataframe(
                resumo_ord.rename(columns={
                    'mes_referencia':  'Mês', 'total_relatorios': 'Relatórios',
                    'total_horas':     'Total Horas', 'total_estudos': 'Total Estudos',
                    'observacao':      'Observação',
                }),
                use_container_width=True, hide_index=True,
            )

            df_pdf_consolidado = resumo_ord[['mes_referencia','total_relatorios','total_horas','total_estudos']].copy()
            df_pdf_consolidado = df_pdf_consolidado.rename(columns={
                'total_horas': 'horas', 'total_estudos': 'estudos_biblicos',
            })
            df_pdf_consolidado['observacoes'] = df_pdf_consolidado['total_relatorios'].apply(
                lambda n: f"{int(n)} relat."
            )
            df_pdf_consolidado['cat_oficial'] = cat_sel

            pdf_c = gerar_pdf_padrao_s21(f"CONSOLIDADO {cat_sel}", cat_sel, df_pdf_consolidado)
            if pdf_c:
                st.download_button(
                    f"📥 Baixar Cartão Consolidado — {cat_sel}",
                    pdf_c, f"S21_Consolidado_{cat_sel}.pdf",
                    use_container_width=True,
                )
        else:
            st.info("Sem dados para esta categoria.")

    # ---- Sub-aba 3: Registro de Assistência (CORRIGIDO) ----
    with c3_tab:
        db = inicializar_db()
        # Chama a função nova reformulada (layout S-88-T original lado a lado)
        render_tab_assistencia(db, congregacao_id="parque_alianca")
# =============================================================
# 16. ABA: ANÚNCIOS
# =============================================================
def aba_anuncios():
    sub_an = st.tabs(["✏️ Nova Postagem", "🗂️ Gerenciar Postagens"])

    with sub_an[0]:
        tipo = st.radio(
            "Tipo",
            ["📝 Texto / Markdown", "🖼️ Imagem (JPEG/PNG)", "📅 Agenda de Reunião"],
            horizontal=True
        )

        if tipo == "📝 Texto / Markdown":
            titulo_txt  = st.text_input("Título (opcional)")
            conteudo_md = st.text_area("Conteúdo", height=200)
            if conteudo_md:
                with st.expander("Pré-visualização"):
                    st.markdown(conteudo_md)
            if st.button("📤 Publicar", type="primary", use_container_width=True):
                if conteudo_md.strip():
                    salvar_anuncio({"tipo": "texto", "titulo": titulo_txt or "Anúncio",
                                    "conteudo_html": conteudo_md, "renderizar_markdown": True})
                    st.success("✅ Publicado!")
                    st.rerun()
                else:
                    st.error("Conteúdo vazio.")

        elif tipo == "🖼️ Imagem (JPEG/PNG)":
            titulo_img = st.text_input("Legenda (opcional)")
            arquivo    = st.file_uploader("Imagem", type=["jpg","jpeg","png"])
            if arquivo:
                st.image(arquivo, use_column_width=True)
                if st.button("📤 Publicar Imagem", type="primary", use_container_width=True):
                    img_bytes = arquivo.read()
                    mime  = "image/png" if arquivo.name.endswith(".png") else "image/jpeg"
                    b64   = base64.b64encode(img_bytes).decode("utf-8")
                    html_img = (f'<div style="text-align:center;padding:10px;">'
                                f'<img src="data:{mime};base64,{b64}" '
                                f'style="max-width:100%;border-radius:8px;" />'
                                + (f'<p style="margin-top:8px;color:#555;">{titulo_img}</p>'
                                   if titulo_img else "") + '</div>')
                    salvar_anuncio({"tipo": "imagem", "titulo": titulo_img or arquivo.name,
                                    "conteudo_html": html_img, "renderizar_markdown": False})
                    st.success("✅ Imagem publicada!")
                    st.rerun()

        elif tipo == "📅 Agenda de Reunião":
            col_a, col_b = st.columns(2)
            data_texto = col_a.text_input("Período", placeholder="18-24 DE MAIO")
            escritura  = col_b.text_input("Escritura", placeholder="ISAÍAS 62-64")

            col_c, col_d, col_e = st.columns(3)
            cant_ab   = col_c.text_input("Cântico Abertura", placeholder="44")
            cant_meio = col_d.text_input("Cântico NVC",      placeholder="115")
            cant_fin  = col_e.text_input("Cântico Final",    placeholder="151")

            st.markdown("---")
            st.markdown('<div style="background:#1a3566;color:white;padding:7px 12px;'
                        'border-radius:5px;font-weight:bold;margin-bottom:6px;">'
                        'TESOUROS DA PALAVRA DE DEUS</div>', unsafe_allow_html=True)
            n_tes = st.number_input("Nº itens", 1, 6, 3, key="n_tes")
            tesouros = []
            for i in range(int(n_tes)):
                c1, c2 = st.columns([4, 1])
                t     = c1.text_input(f"Item {i+1}", key=f"tes_t_{i}",
                                      label_visibility="collapsed", placeholder=f"Item {i+1}")
                d_dur = c2.text_input("Dur.", key=f"tes_d_{i}",
                                      label_visibility="collapsed", placeholder="10 min")
                tesouros.append({"num": i + 1, "titulo": t, "duracao": d_dur})

            st.markdown("---")
            st.markdown('<div style="background:#8a6200;color:white;padding:7px 12px;'
                        'border-radius:5px;font-weight:bold;margin-bottom:6px;">'
                        'FAÇA SEU MELHOR NO MINISTÉRIO</div>', unsafe_allow_html=True)
            n_min = st.number_input("Nº itens", 1, 6, 3, key="n_min")
            ministerio = []
            base_min = int(n_tes)
            for i in range(int(n_min)):
                c1, c2 = st.columns([4, 1])
                t     = c1.text_input(f"Item {base_min+i+1}", key=f"min_t_{i}",
                                      label_visibility="collapsed", placeholder=f"Item {base_min+i+1}")
                d_dur = c2.text_input("Dur.", key=f"min_d_{i}",
                                      label_visibility="collapsed", placeholder="")
                ministerio.append({"num": base_min + i + 1, "titulo": t, "duracao": d_dur})

            st.markdown("---")
            st.markdown('<div style="background:#cc0000;color:white;padding:7px 12px;'
                        'border-radius:5px;font-weight:bold;margin-bottom:6px;">'
                        'NOSSA VIDA CRISTÃ</div>', unsafe_allow_html=True)
            n_nvc = st.number_input("Nº itens", 1, 10, 2, key="n_nvc")
            vida_crista = []
            base_nvc = int(n_tes) + int(n_min)
            for i in range(int(n_nvc)):
                c1, c2 = st.columns([4, 1])
                t     = c1.text_input(f"Item {base_nvc+i+1}", key=f"nvc_t_{i}",
                                      label_visibility="collapsed", placeholder=f"Item {base_nvc+i+1}")
                d_dur = c2.text_input("Dur.", key=f"nvc_d_{i}",
                                      label_visibility="collapsed", placeholder="")
                vida_crista.append({"num": base_nvc + i + 1, "titulo": t, "duracao": d_dur})

            st.markdown("---")
            agenda_dados = {
                "data_texto": data_texto, "escritura": escritura,
                "cantico_abertura": cant_ab, "cantico_meio": cant_meio,
                "cantico_final":    cant_fin,
                "tesouros": tesouros, "ministerio": ministerio, "vida_crista": vida_crista,
            }
            col_prev, col_pub = st.columns(2)
            with col_prev:
                if st.button("👁 Pré-visualizar", use_container_width=True):
                    st.markdown(gerar_html_agenda(agenda_dados), unsafe_allow_html=True)
            with col_pub:
                if st.button("📤 Publicar Agenda", use_container_width=True, type="primary"):
                    if not data_texto.strip():
                        st.error("Informe o período.")
                    else:
                        salvar_anuncio({
                            "tipo": "agenda", "titulo": data_texto,
                            "conteudo_html": gerar_html_agenda(agenda_dados),
                            "renderizar_markdown": False,
                            "dados_agenda": agenda_dados,
                        })
                        st.success(f"✅ Agenda '{data_texto}' publicada!")
                        st.rerun()

    with sub_an[1]:
        anuncios = carregar_anuncios()
        if not anuncios:
            st.info("Nenhuma postagem encontrada.")
        else:
            st.caption(f"{len(anuncios)} postagem(ns) · mais recente primeiro")
            for a in anuncios:
                tipo_icon = {"texto": "📝", "imagem": "🖼️", "agenda": "📅"}.get(a.get("tipo",""), "📌")
                ts       = a.get("data_postagem")
                data_str = ts.strftime("%d/%m/%Y %H:%M") if hasattr(ts, "strftime") else "–"
                with st.expander(f"{tipo_icon} {a.get('titulo','Sem título')}  ·  {data_str}"):
                    if a.get("renderizar_markdown"):
                        st.markdown(a.get("conteudo_html",""), unsafe_allow_html=False)
                    else:
                        st.markdown(a.get("conteudo_html",""), unsafe_allow_html=True)
                    st.markdown("---")
                    if st.button("🗑 Deletar", key=f"del_an_{a['id']}", type="secondary"):
                        deletar_anuncio(a["id"])


# =============================================================
# 17. ABA: CONFIGURAÇÃO
# =============================================================
def aba_configuracao(df, df_ok, df_mes, mes_sel, membros_db):
    sub_cfg = st.tabs(["✏️ EDITAR RELATÓRIOS", "👥 GERENCIAR MEMBROS", "➕ NOVO MEMBRO"])

    # ---- Sub-aba: Editar Relatórios ----
    with sub_cfg[0]:
        st.markdown(f"#### Relatórios Identificados — {mes_sel}")
        if not df.empty:
            df_ok_mes = df[
                (df['mes_referencia'] == mes_sel) &
                (df['status_validacao'] == "IDENTIFICADO")
            ]
            if df_ok_mes.empty:
                st.info("Nenhum relatório identificado neste mês.")
            else:
                for _, r in df_ok_mes.sort_values('nome_oficial').iterrows():
                    with st.expander(f"📝 {r['nome_oficial']} — {int(r['horas'])}h"):
                        ce1, ce2, ce3 = st.columns([2, 1, 1])
                        idx_cat = (categorias_lista.index(r['cat_oficial'])
                                   if r['cat_oficial'] in categorias_lista else 0)
                        nova_cat = ce1.selectbox("Categoria", categorias_lista,
                                                  index=idx_cat, key=f"e_c_{r['id']}")
                        novas_h  = ce2.number_input("Horas",   value=int(r['horas']),
                                                    key=f"e_h_{r['id']}")
                        novos_e  = ce3.number_input("Estudos", value=int(r['estudos_biblicos']),
                                                    key=f"e_e_{r['id']}")

                        col_save, col_del = st.columns(2)
                        with col_save:
                            if st.button("💾 Salvar", key=f"s_b_{r['id']}",
                                         type="primary", use_container_width=True):
                                try:
                                    inicializar_db().collection("relatorios_parque_alianca") \
                                        .document(r['id']).update({
                                            "horas": novas_h, "estudos_biblicos": novos_e,
                                            "categoria_mes": nova_cat,
                                        })
                                    carregar_relatorios_cached.clear()
                                    st.toast("💾 Alterações salvas com sucesso para este mês!")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Erro ao salvar alterações: {e}")

                        with col_del:
                            with st.popover("🗑️ Deletar", use_container_width=True):
                                st.error("Atenção! Ação irreversível.")
                                st.write(f"Deseja apagar definitivamente o relatório de "
                                         f"**{r['nome_oficial']}** deste mês?")
                                if st.button("Sim, Excluir", key=f"conf_del_{r['id']}",
                                             type="primary", use_container_width=True):
                                    deletar_relatorio(r['id'])

    # ---- Sub-aba: Gerenciar Membros ----
    with sub_cfg[1]:
        st.markdown("#### 👥 Gerenciar Membros")
        st.caption("Categoria aqui é a FONTE DA VERDADE para todos os relatórios.")

        tab_ativos, tab_inativos = st.tabs(["👥 Membros Ativos", "💤 Membros Inativos"])

        def renderizar_formulario_membro(nome):
            m        = membros_db[nome]
            cat_icon = {"PUBLICADOR": "👤", "PIONEIRO AUXILIAR": "💎",
                        "PIONEIRO REGULAR": "⭐"}.get(m.get('categoria',''), "👤")

            with st.expander(f"{cat_icon} **{nome}** — {m.get('categoria','PUBLICADOR')}"):
                col_a, col_b = st.columns(2)

                with col_a:
                    st.markdown("##### 📋 Dados Pessoais")
                    cat_gravada = m.get('categoria', 'PUBLICADOR')
                    if cat_gravada not in categorias_lista:
                        cat_gravada = 'PUBLICADOR'
                    nova_cat = st.selectbox("Categoria de Serviço", categorias_lista,
                                             index=categorias_lista.index(cat_gravada),
                                             key=f"cat_{nome}")
                    data_nasc = st.text_input("📅 Nascimento", value=m.get('data_nascimento',''),
                                               placeholder="DD/MM/AAAA", key=f"nasc_{nome}")
                    data_bat  = st.text_input("🕊️ Batismo",   value=m.get('data_batismo',''),
                                               placeholder="DD/MM/AAAA", key=f"bat_{nome}")
                    tel_emer  = st.text_input("📞 Tel. Emergência",
                                               value=m.get('telefone_emergencia',''),
                                               placeholder="(XX) XXXXX-XXXX", key=f"tel_{nome}")

                with col_b:
                    st.markdown("##### 🏷️ Classificação & Cargo")
                    gen_val  = m.get('genero','')
                    nova_gen = st.selectbox("Gênero", _GENEROS,
                                             index=_GENEROS.index(gen_val) if gen_val in _GENEROS else 0,
                                             key=f"gen_{nome}")
                    cls_val  = m.get('classe','')
                    nova_cls = st.selectbox("Classe", _CLASSES,
                                             index=_CLASSES.index(cls_val) if cls_val in _CLASSES else 0,
                                             key=f"cls_{nome}")
                    status_atual = m.get('status', 'Ativo')
                    novo_status  = st.selectbox("Status", _STATUS_OPCOES,
                                                 index=_STATUS_OPCOES.index(status_atual) if status_atual in _STATUS_OPCOES else 0,
                                                 key=f"status_{nome}")
                    cargos_atuais = cargos_para_lista(m.get('cargo',''))
                    st.markdown("**Cargo(s)**")
                    novos_cargos = []
                    for cargo_op in _CARGOS_LISTA:
                        if st.checkbox(cargo_op, value=(cargo_op in cargos_atuais),
                                       key=f"cgo_{nome}_{cargo_op}"):
                            novos_cargos.append(cargo_op)

                st.divider()
                col_save_m, col_del_m = st.columns([3, 1])

                with col_save_m:
                    if st.button("💾 Salvar Alterações", key=f"save_{nome}",
                                 use_container_width=True, type="primary"):
                        extra = {
                            "data_nascimento":     data_nasc,
                            "data_batismo":        data_bat,
                            "telefone_emergencia": tel_emer,
                            "genero":              nova_gen,
                            "classe":              nova_cls,
                            "cargo":               novos_cargos,
                            "status":              novo_status,
                        }
                        atualizar_membro(nome, nova_cat, extra=extra)
                        st.toast(f"✅ {nome} atualizado!")
                        st.rerun()

                with col_del_m:
                    with st.popover("🗑️ Deletar", use_container_width=True):
                        st.error("⚠️ Ação irreversível!")
                        st.write(f"Remove **{nome}** permanentemente do banco de dados.")
                        if st.button(f"Sim, excluir {nome.split()[0]}", key=f"conf_del_m_{nome}",
                                     type="primary", use_container_width=True):
                            deletar_membro(nome)

        membros_ordenados = sorted(membros_db.keys())

        with tab_ativos:
            ativos = [n for n in membros_ordenados if membros_db[n].get('status', 'Ativo') == 'Ativo']
            if ativos:
                for nome in ativos:
                    renderizar_formulario_membro(nome)
            else:
                st.info("Nenhum membro ativo cadastrado.")

        with tab_inativos:
            inativos = [n for n in membros_ordenados if membros_db[n].get('status', 'Ativo') == 'Inativo']
            if inativos:
                for nome in inativos:
                    renderizar_formulario_membro(nome)
            else:
                st.info("Nenhum membro inativo.")

    # ---- Sub-aba: Novo Membro ----
    with sub_cfg[2]:
        st.markdown("#### ➕ Cadastrar Novo Membro")
        with st.form("novo_membro", clear_on_submit=True):
            st.markdown("##### Dados Obrigatórios")
            c1, c2 = st.columns(2)
            nm = c1.text_input("Nome Completo *")
            ct = c2.selectbox("Categoria *", categorias_lista)

            st.markdown("##### Dados do Cartão S-21")
            c3, c4 = st.columns(2)
            data_nasc_n = c3.text_input("📅 Nascimento", placeholder="DD/MM/AAAA")
            data_bat_n  = c4.text_input("🕊️ Batismo",    placeholder="DD/MM/AAAA")

            c5, c6 = st.columns(2)
            gen_n = c5.selectbox("Gênero", ["","Masculino","Feminino"])
            cls_n = c6.selectbox("Classe", ["","Outras ovelhas","Ungido"])

            st.markdown("**Cargo(s)**")
            cargos_novos_form = []
            cols_form = st.columns(len(_CARGOS_LISTA))
            for idx_c, cargo_op in enumerate(_CARGOS_LISTA):
                if cols_form[idx_c].checkbox(cargo_op, key=f"new_cgo_{cargo_op}"):
                    cargos_novos_form.append(cargo_op)

            tel_n = st.text_input("📞 Telefone de Emergência", placeholder="(XX) XXXXX-XXXX")

            if st.form_submit_button("➕ Adicionar Membro", use_container_width=True, type="primary"):
                if nm.strip():
                    extra_n = {
                        "data_nascimento":     data_nasc_n,
                        "data_batismo":        data_bat_n,
                        "telefone_emergencia": tel_n,
                        "genero":              gen_n,
                        "classe":              cls_n,
                        "cargo":               cargos_novos_form,
                        "status":              "Ativo",
                    }
                    atualizar_membro(nm.strip(), ct, novo=True, extra=extra_n)
                    st.success(f"✅ {nm.strip()} adicionado!")
                    st.rerun()
                else:
                    st.error("Informe o nome completo.")


# =============================================================
# 18. MÓDULO: PASSAGENS (VGP)
# =============================================================
# Código originalmente em passagens.py, incorporado aqui para
# manter tudo em um único arquivo (mesmo padrão do módulo de
# Assistência acima).
#
# AJUSTES feitos ao colar aqui dentro do main.py:
#   1) Removido o st.set_page_config() e o bloco de <style> do
#      topo do arquivo original — o app já tem os dois (seção 2),
#      e chamar set_page_config() duas vezes quebra o Streamlit.
#   2) O passagens.py originalmente usava um Firestore separado
#      (projeto "bancowendley"), o que exigiria liberar permissão
#      de IAM para a service account nesse projeto. Por decisão
#      do usuário, o módulo passou a usar a MESMA conexão/projeto
#      do app principal (inicializar_db() → "wendleydesenvolvimento").
#      Os dados que já existiam no "bancowendley" NÃO aparecem
#      aqui (é um banco físico diferente) — a aba Passagens
#      começa vazia neste projeto.
# =============================================================

CAPACIDADE = 46


def atualizar_cadastro_central(dados_pax):
    db = inicializar_db()
    if db:
        pax_id = dados_pax['nome'].lower().replace(" ", "")
        db.collection("cadastro_geral").document(pax_id).set({
            "nome": dados_pax['nome'], "rg": dados_pax.get('rg', ""),
            "cpf": dados_pax.get('cpf', ""), "grupo": dados_pax.get('grupo', "Geral"),
            "ultima_atualizacao": datetime.now()
        }, merge=True)


def buscar_pessoa_central(nome_pesquisa):
    db = inicializar_db()
    if not db or not nome_pesquisa: return None
    nome_busca = nome_pesquisa.lower().strip()
    for doc in db.collection("cadastro_geral").stream():
        dados = doc.to_dict()
        if nome_busca in dados.get('nome', '').lower():
            return dados
    return None


def criar_evento(nome, datas, valor_passagem):
    db = inicializar_db()
    if db:
        id_evento = f"{nome.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}"
        db.collection("eventos").document(id_evento).set({
            "nome": nome, "datas": datas, "valor": valor_passagem,
            "status": "ativo", "criado_em": datetime.now(),
            "frotas": {dia: 1 for dia in datas}
        })
        return id_evento


def adicionar_novo_onibus(id_evento, dia):
    db = inicializar_db()
    if db:
        doc_ref = db.collection("eventos").document(id_evento)
        evento  = doc_ref.get().to_dict()
        frotas  = evento.get('frotas', {d: 1 for d in evento['datas']})
        frotas[dia] = frotas.get(dia, 1) + 1
        doc_ref.update({"frotas": frotas})


def salvar_passageiro(id_evento, dados_pax):
    db = inicializar_db()
    if db:
        sufixo = dados_pax['rg'] if dados_pax.get('rg') else "reserva"
        pax_id = f"{dados_pax['nome']}_{sufixo}".lower().replace(" ", "")
        if 'embarcou' not in dados_pax: dados_pax['embarcou'] = False
        db.collection("eventos").document(id_evento).collection("passageiros").document(pax_id).set(dados_pax)
        atualizar_cadastro_central(dados_pax)


def atualizar_embarque(id_evento, pax, status):
    db = inicializar_db()
    if db:
        sufixo = pax['rg'] if pax.get('rg') else "reserva"
        pax_id = f"{pax['nome']}_{sufixo}".lower().replace(" ", "")
        db.collection("eventos").document(id_evento).collection("passageiros").document(pax_id).update({"embarcou": status})


def deletar_passageiro(id_evento, nome, rg):
    db = inicializar_db()
    if db:
        sufixo = rg if rg else "reserva"
        pax_id = f"{nome}_{sufixo}".lower().replace(" ", "")
        db.collection("eventos").document(id_evento).collection("passageiros").document(pax_id).delete()


def carregar_passageiros(id_evento):
    db = inicializar_db()
    return [p.to_dict() for p in db.collection("eventos").document(id_evento).collection("passageiros").stream()]


def carregar_eventos():
    db = inicializar_db()
    if not db: return {}
    return {doc.id: doc.to_dict() for doc in db.collection("eventos").where("status", "==", "ativo").stream()}


@st.dialog("Gerenciar Reserva")
def gerenciar_pax_dialog(pax, id_evento, evento_atual):
    st.markdown("### 👤 " + pax['nome'])
    total_devido    = pax.get('valor_total', len(pax.get('dias_onibus', [])) * evento_atual['valor'])
    pago_atualmente = pax.get('valor_pago', 0.0)
    c1, c2 = st.columns(2)
    c1.metric("Total da Passagem", "R$ %.2f" % total_devido)
    c2.metric("Saldo Pendente",    "R$ %.2f" % (total_devido - pago_atualmente), delta_color="inverse")

    with st.form("edit_pax_final"):
        nome = st.text_input("Nome", value=pax['nome'])
        cc1, cc2 = st.columns(2)
        rg  = cc1.text_input("RG",  value=pax.get('rg', ""))
        cpf = cc2.text_input("CPF", value=pax.get('cpf', ""))
        grupos = ["Rosas", "Engenho", "Cohab", "Geral"]
        g_atual = pax.get('grupo', 'Geral')
        grupo = st.selectbox("Grupo", grupos, index=grupos.index(g_atual) if g_atual in grupos else 3)

        st.divider()
        st.markdown("**💰 Registrar Recebimento**")
        cr1, cr2, cr3 = st.columns(3)
        valor_recebido = cr1.number_input("Recebido agora", min_value=0.0, value=0.0, step=5.0)
        valor_entregue = cr2.number_input("Troco entregue", min_value=0.0, value=0.0)
        if valor_entregue > 0 and valor_recebido > 0:
            cr3.success("Troco: R$ %.2f" % max(valor_entregue - valor_recebido, 0))

        st.divider()
        st.markdown("**🗓 Viagens**")
        novas_viagens  = []
        viagens_atuais = {v['dia']: v['bus'] for v in pax.get('dias_onibus', [])}
        for dia in evento_atual['datas']:
            cd1, cd2 = st.columns([1, 2])
            if cd1.checkbox(dia, value=dia in viagens_atuais, key="edit_chk_" + dia):
                n_frotas    = evento_atual.get('frotas', {}).get(dia, 1)
                bus_default = viagens_atuais.get(dia, 1)
                bus_sel     = cd2.selectbox("Ônibus " + dia, range(1, n_frotas + 1),
                                            index=min(bus_default - 1, n_frotas - 1),
                                            key="edit_sel_" + dia)
                novas_viagens.append({"dia": dia, "bus": bus_sel})

        st.divider()
        novo_total_pago = pago_atualmente + valor_recebido
        pago     = st.toggle("💰 Pagamento quitado", value=pax.get('pago', False) or (novo_total_pago >= total_devido))
        embarque = st.toggle("🚌 Embarcou",           value=pax.get('embarcou', False))

        cb1, cb2 = st.columns(2)
        if cb1.form_submit_button("💾 Salvar", use_container_width=True, type="primary"):
            if nome != pax['nome'] or rg != pax.get('rg', ""):
                deletar_passageiro(id_evento, pax['nome'], pax.get('rg', ""))
            pax.update({"nome": nome, "rg": rg, "cpf": cpf, "grupo": grupo,
                        "dias_onibus": novas_viagens, "pago": pago, "embarcou": embarque,
                        "valor_total": evento_atual['valor'] * len(novas_viagens),
                        "valor_pago": novo_total_pago})
            salvar_passageiro(id_evento, pax)
            st.rerun()
        if cb2.form_submit_button("🗑️ Excluir", use_container_width=True):
            deletar_passageiro(id_evento, pax['nome'], pax.get('rg', ""))
            st.rerun()


def renderizar_cabecalho_passagens(evento, df, id_sel):
    total      = len(df) if not df.empty else 0
    pagos      = int(df['pago'].sum())         if not df.empty and 'pago'      in df.columns else 0
    pendente   = total - pagos
    arrecadado = float(df['valor_pago'].sum()) if not df.empty and 'valor_pago' in df.columns else 0.0
    a_receber  = float((df['valor_total'].fillna(0) - df['valor_pago'].fillna(0)).clip(lower=0).sum()) \
                 if not df.empty and 'valor_total' in df.columns else 0.0
    pct        = round((pagos / total) * 100) if total else 0
    datas_str  = ", ".join(evento.get("datas", []))
    nome_ev    = evento.get('nome', '')

    # ---- Montar HTML dos cards de frota ----
    frotas_html  = ""
    needs_add    = {}
    for dia in evento.get('datas', []):
        n_frotas = evento.get('frotas', {}).get(dia, 1)
        for b in range(1, n_frotas + 1):
            qtd = 0
            if not df.empty:
                for _, p in df.iterrows():
                    for v in (p.get('dias_onibus') or []):
                        if v.get('dia') == dia and v.get('bus') == b:
                            qtd += 1
            perc     = min(round((qtd / CAPACIDADE) * 100), 100)
            cor      = "#f87171" if qtd >= CAPACIDADE else ("#C9A227" if perc > 80 else "#4ade80")
            lotado   = "<div style='font-size:0.6rem;color:#f87171;font-weight:700;margin-top:3px;'>🔴 Lotado</div>" \
                       if qtd >= CAPACIDADE else ""
            frotas_html += (
                "<div style='background:rgba(255,255,255,0.09);border:1px solid rgba(255,255,255,0.15);"
                "border-radius:10px;padding:11px 13px;min-width:130px;flex:1;'>"
                "<div style='display:flex;justify-content:space-between;align-items:baseline;margin-bottom:7px;'>"
                "<span style='font-size:0.72rem;font-weight:700;color:white;'>" + dia + " · Ônibus " + str(b) + "</span>"
                "<span style='font-size:0.65rem;font-weight:600;color:rgba(255,255,255,0.5);'>" + str(perc) + "%</span>"
                "</div>"
                "<div style='background:rgba(255,255,255,0.14);border-radius:4px;height:6px;overflow:hidden;margin-bottom:5px;'>"
                "<div style='width:" + str(perc) + "%;height:100%;background:" + cor + ";border-radius:4px;'></div>"
                "</div>"
                "<div style='font-size:0.63rem;color:rgba(255,255,255,0.42);'>" + str(qtd) + " / " + str(CAPACIDADE) + " passageiros</div>"
                + lotado +
                "</div>"
            )
            if qtd >= CAPACIDADE and b == n_frotas:
                needs_add[dia] = b + 1

    # ---- Montar HTML dos KPIs ----
    def kpi(lbl, val, sub, cor="white"):
        return (
            "<div style='background:rgba(255,255,255,0.11);border:1px solid rgba(255,255,255,0.18);"
            "border-radius:11px;padding:12px 14px;flex:1;min-width:110px;'>"
            "<div style='font-size:0.62rem;color:rgba(255,255,255,0.55);text-transform:uppercase;"
            "letter-spacing:0.09em;font-weight:700;margin-bottom:5px;'>" + lbl + "</div>"
            "<div style='font-size:1.4rem;font-weight:700;color:" + cor + ";line-height:1;'>" + str(val) + "</div>"
            "<div style='font-size:0.65rem;color:rgba(255,255,255,0.4);margin-top:3px;'>" + sub + "</div>"
            "</div>"
        )

    kpis_html = (
        kpi("Reservas",   total,                             "passageiros")
      + kpi("Pagos",      pagos,                             str(pct) + "% confirmados", "#f5e3a8")
      + kpi("Pendentes",  pendente,                          "aguardando",                "#ffd166")
      + kpi("Arrecadado", "R$ {:,.0f}".format(arrecadado),  "recebido",                  "#f5e3a8")
      + kpi("A Receber",  "R$ {:,.0f}".format(a_receber),   "em aberto",                 "#ffd166")
    )

    n_frotas_total       = sum(evento.get('frotas', {}).get(d, 1) for d in evento.get('datas', []))
    linhas_frota_mobile  = n_frotas_total
    altura = (
        72
        + 3 * 78
        + 50
        + linhas_frota_mobile * 82
        + 56
    )

    # Cabeçalho do módulo em preto+dourado — combina com a barra
    # superior do resto do app, em vez do azul original.
    html = (
        "<link href='https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap' rel='stylesheet'>"
        "<style>* { box-sizing:border-box; } body { background:transparent; overflow:hidden; margin:0; }</style>"
        "<div id='root' style='font-family:Inter,sans-serif;"
        "background:linear-gradient(135deg,#161514 0%,#2a2620 100%);"
        "border:1px solid #C9A227;border-radius:16px;padding:24px 24px 20px;color:white;'>"

        "<div style='font-size:1.5rem;font-weight:700;letter-spacing:-0.5px;color:#F0D98C;'>"
        "🕊️ " + nome_ev +
        "</div>"
        "<div style='font-size:0.8rem;color:rgba(255,255,255,0.6);margin-top:4px;font-weight:400;'>"
        "Controle de Passagens · " + datas_str +
        "</div>"

        "<div style='display:grid;grid-template-columns:repeat(auto-fill,minmax(120px,1fr));"
        "gap:10px;margin-top:16px;'>"
        + kpis_html +
        "</div>"

        "<div style='border-top:1px solid rgba(201,162,39,0.35);margin:16px 0 14px;'></div>"

        "<div style='font-size:0.6rem;font-weight:700;letter-spacing:0.1em;text-transform:uppercase;"
        "color:rgba(255,255,255,0.45);margin-bottom:10px;'>Ocupação por Frota</div>"
        "<div style='display:grid;grid-template-columns:repeat(auto-fill,minmax(110px,1fr));gap:8px;'>"
        + frotas_html +
        "</div>"

        "</div>"

        "<script>"
        "function reportH(){"
        "  var h=document.getElementById('root').getBoundingClientRect().height+8;"
        "  window.parent.postMessage({type:'streamlit:setFrameHeight',height:h},'*');"
        "}"
        "window.addEventListener('load', function(){ setTimeout(reportH, 50); });"
        "new ResizeObserver(function(){ setTimeout(reportH, 50); })"
        "  .observe(document.getElementById('root'));"
        "</script>"
    )

    components.html(html, height=altura, scrolling=False)

    if needs_add:
        cols = st.columns(len(needs_add))
        for idx, (dia, prox) in enumerate(needs_add.items()):
            with cols[idx]:
                if st.button("➕ Adicionar Ônibus " + str(prox) + " — " + dia, key="hdr_add_" + dia):
                    adicionar_novo_onibus(id_sel, dia)
                    st.rerun()


def exibir_modulo_passagens():
    eventos_ativos = carregar_eventos()

    if not eventos_ativos:
        components.html(
            "<link href='https://fonts.googleapis.com/css2?family=Inter:wght@700&display=swap' rel='stylesheet'>"
            "<div style='font-family:Inter,sans-serif;background:linear-gradient(135deg,#161514 0%,#2a2620 100%);"
            "border:1px solid #C9A227;border-radius:16px;padding:24px;color:white;'>"
            "<div style='font-size:1.5rem;font-weight:700;color:#F0D98C;'>🕊️ VGP Passagens</div>"
            "<div style='font-size:0.82rem;color:rgba(255,255,255,0.6);margin-top:4px;'>"
            "Nenhum evento ativo — crie o primeiro abaixo</div></div>",
            height=110
        )
        with st.form("criar_evento_inicial"):
            st.subheader("Novo Evento")
            n_ev = st.text_input("Nome do Evento (ex: Assembleia Março)")
            v_ev = st.number_input("Valor da Passagem (R$)", min_value=0.0, value=50.0, step=5.0)
            d_ev = st.multiselect("Dias de Operação", ["Sexta", "Sábado", "Domingo"])
            if st.form_submit_button("🚀 Criar Evento", type="primary"):
                if n_ev and d_ev:
                    criar_evento(n_ev, d_ev, v_ev)
                    st.rerun()
                else:
                    st.error("Informe o nome e ao menos um dia.")
        return

    c1, c2 = st.columns([4, 1])
    with c2:
        id_sel = st.selectbox("", list(eventos_ativos.keys()),
                              format_func=lambda x: eventos_ativos[x]['nome'],
                              label_visibility="collapsed", key="passagens_evento_sel")

    evento    = eventos_ativos[id_sel]
    pax_lista = carregar_passageiros(id_sel)
    df        = pd.DataFrame(pax_lista)

    if not df.empty:
        for col, default in [('grupo','Geral'),('pago',False),('valor_pago',0.0),
                              ('valor_total',0.0),('embarcou',False)]:
            if col not in df.columns: df[col] = default
            df[col] = df[col].fillna(default)

    renderizar_cabecalho_passagens(evento, df, id_sel)

    tab_reserva, tab_chamada, tab_ajustes = st.tabs([
        "📝 Reserva & Pagamentos",
        "🚌 Chamada de Embarque",
        "⚙️ Ajustes"
    ])

    # ---- ABA 1: RESERVA + PENDENTES ----
    with tab_reserva:
        col_form, col_pend = st.columns([1, 1], gap="large")

        with col_form:
            st.markdown("**Nova Reserva**")
            busca_nome = st.text_input("🔍 Buscar cadastro existente", placeholder="Digite parte do nome...")
            mestre = buscar_pessoa_central(busca_nome) if busca_nome else None
            if mestre:
                st.success("✅ Cadastro encontrado: **" + mestre['nome'] + "**")

            with st.form("reserva_form", clear_on_submit=True):
                nome_f  = st.text_input("Nome Completo *", value=mestre['nome'] if mestre else busca_nome)
                ci1, ci2 = st.columns(2)
                rg_f  = ci1.text_input("RG",  value=mestre.get('rg',  '') if mestre else "")
                cpf_f = ci2.text_input("CPF", value=mestre.get('cpf', '') if mestre else "")
                grupo_f = st.selectbox("Grupo / Localização", ["Rosas", "Engenho", "Cohab", "Geral"])
                st.markdown("**Viagens:**")
                viagens = []
                for dia in evento['datas']:
                    cv1, cv2 = st.columns([1, 2])
                    if cv1.checkbox(dia, key="f_res_" + dia):
                        f_dia = evento.get('frotas', {}).get(dia, 1)
                        b_sel = cv2.selectbox("Ônibus " + dia, range(1, f_dia + 1), key="f_bus_" + dia)
                        viagens.append({"dia": dia, "bus": b_sel})
                pago_f = st.toggle("Pagamento confirmado neste ato")
                if st.form_submit_button("✅ Confirmar Reserva", type="primary", use_container_width=True):
                    if nome_f and viagens:
                        vt = evento['valor'] * len(viagens)
                        salvar_passageiro(id_sel, {
                            "nome": nome_f, "rg": rg_f, "cpf": cpf_f, "grupo": grupo_f,
                            "dias_onibus": viagens, "pago": pago_f, "embarcou": False,
                            "valor_total": vt, "valor_pago": vt if pago_f else 0.0
                        })
                        st.success("Reserva gravada com sucesso!")
                        st.rerun()
                    else:
                        st.error("Informe o nome e selecione ao menos um dia.")

        with col_pend:
            st.markdown("**Pagamentos Pendentes**")
            if not df.empty:
                pendentes = df[df['pago'] == False].sort_values('nome')
                if pendentes.empty:
                    st.markdown(
                        "<div style='text-align:center;padding:40px 0;color:#94a3b8;'>"
                        "<div style='font-size:2rem;'>✅</div>"
                        "<div style='font-weight:600;margin-top:8px;'>Todos pagos!</div>"
                        "</div>", unsafe_allow_html=True)
                else:
                    total_pend = 0.0
                    for _, r in pendentes.iterrows():
                        v_total = float(r.get('valor_total') or 0) or (len(r.get('dias_onibus') or []) * evento['valor'])
                        v_pago  = float(r.get('valor_pago')  or 0)
                        v_falta = max(v_total - v_pago, 0)
                        total_pend += v_falta
                        grp_tag = r.get('grupo', 'Geral')
                        ci, cb = st.columns([5, 1])
                        with ci:
                            st.markdown(
                                "<div style='background:white;border:1px solid #EFE3B8;"
                                "border-left:4px solid #ef4444;border-radius:10px;"
                                "padding:10px 13px;margin-bottom:7px;"
                                "display:flex;justify-content:space-between;align-items:center;'>"
                                "<div>"
                                "<div style='font-weight:600;font-size:0.87rem;color:#1e293b;'>" + r['nome'] + "</div>"
                                "<div style='font-size:0.74rem;color:#9C8A46;margin-top:2px;'>"
                                "📍 " + grp_tag + " · " + str(len(r.get('dias_onibus') or [])) + " viagem(ns)</div>"
                                "</div>"
                                "<div style='font-weight:700;font-size:0.9rem;color:#ef4444;"
                                "white-space:nowrap;margin-left:8px;'>"
                                "– R$ {:,.2f}".format(v_falta) + "</div>"
                                "</div>", unsafe_allow_html=True)
                        with cb:
                            if st.button("✏️", key="ed_pe_" + r['nome'], help="Editar / Receber pagamento"):
                                gerenciar_pax_dialog(r.to_dict(), id_sel, evento)

                    st.markdown(
                        "<div style='background:#FBF1D4;border:1px solid #E9D48E;border-radius:8px;"
                        "padding:10px 14px;margin-top:10px;font-size:0.85rem;"
                        "display:flex;justify-content:space-between;align-items:center;'>"
                        "<strong>Total em aberto:</strong>"
                        "<span style='font-weight:700;color:#8A6D14;'>R$ {:,.2f}</span>".format(total_pend) +
                        "</div>", unsafe_allow_html=True)
            else:
                st.info("Nenhuma reserva lançada ainda.")

    # ---- ABA 2: CHAMADA ----
    with tab_chamada:
        if df.empty:
            st.info("Nenhuma reserva para exibir.")
        else:
            df_pagos = df[df['pago'] == True].copy()
            if df_pagos.empty:
                st.markdown(
                    "<div style='text-align:center;padding:60px 0;color:#94a3b8;'>"
                    "<div style='font-size:2.5rem;'>🕊️</div>"
                    "<div style='font-weight:600;margin-top:10px;'>Nenhum pagamento confirmado ainda.</div>"
                    "<div style='font-size:0.82rem;margin-top:6px;'>Só passageiros com pagamento quitado aparecem aqui.</div>"
                    "</div>", unsafe_allow_html=True)
            else:
                tot_p  = len(df_pagos)
                emb_t  = int(df_pagos['embarcou'].sum())
                falt_t = tot_p - emb_t
                st.markdown(
                    "<div style='display:flex;gap:10px;margin-bottom:18px;flex-wrap:wrap;'>"
                    "<div style='background:white;border:1px solid #EFE3B8;border-radius:10px;"
                    "padding:12px 18px;flex:1;min-width:110px;'>"
                    "<div style='font-size:0.62rem;color:#9C8A46;text-transform:uppercase;"
                    "letter-spacing:.08em;font-weight:700;'>Confirmados</div>"
                    "<div style='font-size:1.5rem;font-weight:700;color:#1A1A1A;'>" + str(tot_p) + "</div></div>"

                    "<div style='background:white;border:1px solid #EFE3B8;border-left:3px solid #22c55e;"
                    "border-radius:10px;padding:12px 18px;flex:1;min-width:110px;'>"
                    "<div style='font-size:0.62rem;color:#9C8A46;text-transform:uppercase;"
                    "letter-spacing:.08em;font-weight:700;'>Embarcados</div>"
                    "<div style='font-size:1.5rem;font-weight:700;color:#22c55e;'>" + str(emb_t) + "</div></div>"

                    "<div style='background:white;border:1px solid #EFE3B8;border-left:3px solid #C9A227;"
                    "border-radius:10px;padding:12px 18px;flex:1;min-width:110px;'>"
                    "<div style='font-size:0.62rem;color:#9C8A46;text-transform:uppercase;"
                    "letter-spacing:.08em;font-weight:700;'>Aguardando</div>"
                    "<div style='font-size:1.5rem;font-weight:700;color:#B4952E;'>" + str(falt_t) + "</div></div>"
                    "</div>", unsafe_allow_html=True)

                for grp in sorted(df_pagos['grupo'].unique()):
                    df_grp = df_pagos[df_pagos['grupo'] == grp]
                    n_grp  = len(df_grp)
                    e_grp  = int(df_grp['embarcou'].sum())
                    with st.expander("📍 " + grp.upper() + "  —  " + str(e_grp) + "/" + str(n_grp) + " embarcados", expanded=True):
                        cf, co = st.columns(2)
                        with cf:
                            st.markdown("<div style='font-size:0.7rem;font-weight:700;text-transform:uppercase;"
                                        "letter-spacing:.08em;color:#B4952E;margin-bottom:8px;'>⏳ Aguardando</div>",
                                        unsafe_allow_html=True)
                            for _, p in df_grp[df_grp['embarcou'] == False].sort_values('nome').iterrows():
                                cn, cb = st.columns([5, 1])
                                cn.markdown("<div style='font-weight:500;font-size:0.87rem;color:#1e293b;"
                                            "padding:6px 0;border-bottom:1px solid #F1EAD2;'>" + p['nome'] + "</div>",
                                            unsafe_allow_html=True)
                                if cb.button("✅", key="emb_" + grp + "_" + p['nome']):
                                    atualizar_embarque(id_sel, p.to_dict(), True); st.rerun()
                        with co:
                            st.markdown("<div style='font-size:0.7rem;font-weight:700;text-transform:uppercase;"
                                        "letter-spacing:.08em;color:#22c55e;margin-bottom:8px;'>🟢 Embarcados</div>",
                                        unsafe_allow_html=True)
                            for _, p in df_grp[df_grp['embarcou'] == True].sort_values('nome').iterrows():
                                cn, cb = st.columns([5, 1])
                                cn.markdown("<div style='font-weight:500;font-size:0.87rem;color:#94a3b8;"
                                            "text-decoration:line-through;padding:6px 0;"
                                            "border-bottom:1px solid #F1EAD2;'>" + p['nome'] + "</div>",
                                            unsafe_allow_html=True)
                                if cb.button("↩️", key="rem_" + grp + "_" + p['nome']):
                                    atualizar_embarque(id_sel, p.to_dict(), False); st.rerun()

    # ---- ABA 3: AJUSTES ----
    with tab_ajustes:
        ca1, ca2 = st.columns(2)
        with ca1:
            st.markdown("**Novo Evento**")
            with st.form("criar_evento_adj"):
                n_ev = st.text_input("Nome do Evento")
                v_ev = st.number_input("Valor da Passagem (R$)", min_value=0.0, value=50.0, step=5.0)
                d_ev = st.multiselect("Dias de Operação", ["Sexta", "Sábado", "Domingo"])
                if st.form_submit_button("🚀 Criar Evento", type="primary"):
                    if n_ev and d_ev:
                        criar_evento(n_ev, d_ev, v_ev); st.rerun()
            st.divider()
            st.markdown("**Exportar Dados**")
            if not df.empty:
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    df.to_excel(writer, index=False, sheet_name='Passageiros')
                st.download_button("📥 Baixar Excel", output.getvalue(),
                                   "lista_" + id_sel + ".xlsx", use_container_width=True)
        with ca2:
            st.markdown("**Encerrar Evento**")
            with st.container(border=True):
                st.warning("Encerrar **" + evento['nome'] + "** o moverá para o histórico.")
                confirmacao = st.text_input("Digite o nome do evento para confirmar:", placeholder=evento['nome'])
                if st.button("🏁 Arquivar Evento", type="primary", use_container_width=True):
                    if confirmacao.strip().lower() == evento['nome'].strip().lower():
                        inicializar_db().collection("eventos").document(id_sel).update({"status": "finalizado"})
                        st.success("Evento arquivado.")
                        st.rerun()
                    else:
                        st.error("Nome não confere. Tente novamente.")


# =============================================================
# 19. PONTO DE ENTRADA PRINCIPAL
# =============================================================
def main():
    # Verificar autenticação
    if not st.session_state.get("autenticado"):
        tela_login()
        st.stop()

    # Carregar dados
    membros_db        = carregar_membros()
    relatorios_brutos = carregar_relatorios()
    registros_assist  = carregar_assistencia()
    df                = processar_dataframe(relatorios_brutos, membros_db)
    mes_vigente       = obter_mes_vigente_str()

    # Sidebar → retorna o mês selecionado
    mes_sel = renderizar_sidebar(df, mes_vigente)

    # Cabeçalho da página
    col_title, col_mes = st.columns([3, 1])
    with col_title:
        st.markdown("# Parque Aliança")
        st.markdown(
            '<p style="color:#6b7280;font-size:0.85rem;margin-top:-8px;">'
            'Sistema de Gestão · Relatórios & Publicadores</p>',
            unsafe_allow_html=True
        )
    with col_mes:
        st.markdown(f"""
        <div style="text-align:right;margin-top:12px;">
            <div class="mes-badge">
                <span class="mes-dot"></span>{mes_vigente}
            </div>
        </div>""", unsafe_allow_html=True)

    st.markdown("---")

    # DataFrames derivados
    df_mes = df[df['mes_referencia'] == mes_sel] if not df.empty else pd.DataFrame()
    df_ok  = df_mes[df_mes['status_validacao'] == "IDENTIFICADO"] if not df_mes.empty else pd.DataFrame()

    # Abas principais
    tabs = st.tabs([
        "📋  RELATÓRIOS",
        "⚠️  TRIAGEM",
        "📈  CONSOLIDADO",
        "📢  ANÚNCIOS",
        "🚌  PASSAGENS",
        "⚙️  CONFIGURAÇÃO",
    ])

    with tabs[0]:
        aba_relatorios(df_ok, df_mes, mes_sel, membros_db, df)

    with tabs[1]:
        aba_triagem(df_mes, membros_db)

    with tabs[2]:
        aba_consolidado(df, membros_db, mes_vigente, registros_assist)

    with tabs[3]:
        aba_anuncios()

    with tabs[4]:
        exibir_modulo_passagens()

    with tabs[5]:
        aba_configuracao(df, df_ok, df_mes, mes_sel, membros_db)

    # Rodapé
    st.markdown("""
    <div style="text-align:center;padding:2rem 0 0.5rem;
        font-size:0.72rem;color:#374151;letter-spacing:0.05em;">
        v5.2 · Parque Aliança · Sistema de Gestão
    </div>""", unsafe_allow_html=True)


if __name__ == "__main__":
    main()
